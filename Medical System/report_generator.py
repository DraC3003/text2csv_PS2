"""
Report Generator for Medical Test System
Handles Excel and PDF report generation with color coding and formatting.
"""

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import LineChart, Reference
import matplotlib.pyplot as plt
import matplotlib.dates as mdates
from matplotlib.backends.backend_pdf import PdfPages
from datetime import datetime, timedelta
import numpy as np
import os
from typing import List, Dict, Tuple
from database_manager import DatabaseManager

class ReportGenerator:
    def __init__(self, db_manager: DatabaseManager):
        """Initialize report generator with database manager"""
        self.db_manager = db_manager
        
        # Define color schemes for different test result ranges
        self.colors = {
            'normal': PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid'),  # Light green
            'high': PatternFill(start_color='FFB6C1', end_color='FFB6C1', fill_type='solid'),    # Light red
            'low': PatternFill(start_color='87CEEB', end_color='87CEEB', fill_type='solid'),     # Light blue
            'critical_high': PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid'),  # Red
            'critical_low': PatternFill(start_color='0000FF', end_color='0000FF', fill_type='solid'),   # Blue
            'header': PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')   # Light gray
        }
        
        self.fonts = {
            'header': Font(bold=True, size=12),
            'normal': Font(size=10),
            'bold': Font(bold=True, size=10)
        }
    
    def determine_test_status(self, test_value: float, normal_min: float, normal_max: float, 
                             test_name: str = None, age: int = None, gender: str = None) -> str:
        """Determine the status of a test result based on age and gender-adjusted normal ranges"""
        # Get age/gender adjusted ranges if available
        if test_name:
            range_info = self.db_manager.get_age_gender_adjusted_range(test_name, age, gender)
            if range_info['normal_min'] is not None and range_info['normal_max'] is not None:
                normal_min, normal_max = range_info['normal_min'], range_info['normal_max']
                # Store range info for potential status message enhancement
                self._last_range_info = range_info
        
        if normal_min is None or normal_max is None:
            return 'unknown'
        
        # Get critical thresholds from range info if available
        if hasattr(self, '_last_range_info') and self._last_range_info:
            critical_low = self._last_range_info.get('critical_low')
            critical_high = self._last_range_info.get('critical_high')
        else:
            critical_low = critical_high = None
        
        # Convert to float if they exist and are not None
        try:
            if critical_low is not None:
                critical_low = float(critical_low)
            if critical_high is not None:
                critical_high = float(critical_high)
        except (ValueError, TypeError):
            critical_low = critical_high = None
        
        # If no specific critical thresholds, calculate them (30% beyond normal range)
        if critical_low is None or critical_high is None:
            range_width = normal_max - normal_min
            critical_high = normal_max + (range_width * 0.3)
            critical_low = normal_min - (range_width * 0.3)
        
        if test_value >= critical_high:
            return 'critical_high'
        elif test_value <= critical_low:
            return 'critical_low'
        elif test_value > normal_max:
            return 'high'
        elif test_value < normal_min:
            return 'low'
        else:
            return 'normal'
    
    def generate_excel_report(self, patient_id: str, output_path: str) -> bool:
        """Generate a comprehensive Excel report for a patient"""
        try:
            # Get patient information
            patient_info = self.db_manager.get_patient(patient_id)
            if not patient_info:
                raise ValueError(f"Patient {patient_id} not found")
            
            # Get test results
            test_results = self.db_manager.get_patient_test_results(patient_id)
            if not test_results:
                raise ValueError(f"No test results found for patient {patient_id}")
            
            # Create workbook and worksheets
            wb = openpyxl.Workbook()
            
            # Remove default sheet
            wb.remove(wb.active)
            
            # Create patient summary sheet
            self._create_patient_summary_sheet(wb, patient_info, test_results)
            
            # Create detailed results sheet
            self._create_detailed_results_sheet(wb, test_results)
            
            # Create trends sheet if multiple results exist
            if len(test_results) > 1:
                self._create_trends_sheet(wb, test_results)
            
            # Save workbook
            wb.save(output_path)
            return True
            
        except Exception as e:
            raise Exception(f"Failed to generate Excel report: {str(e)}")
    
    def generate_all_patients_excel_report(self, output_path: str) -> bool:
        """Generate a comprehensive Excel report for ALL patients with color coding"""
        try:
            # Get all patients
            all_patients = self.db_manager.get_all_patients()
            if not all_patients:
                raise ValueError("No patients found in database")
            
            # Create workbook
            wb = openpyxl.Workbook()
            wb.remove(wb.active)  # Remove default sheet
            
            # Create summary sheet for all patients
            self._create_all_patients_summary_sheet(wb, all_patients)
            
            # Create detailed sheet for each patient with test results
            for patient in all_patients:
                patient_id = patient[0]
                test_results = self.db_manager.get_patient_test_results(patient_id)
                if test_results:  # Only create sheet if patient has test results
                    self._create_patient_detail_sheet(wb, patient, test_results)
            
            # Create overview statistics sheet
            self._create_statistics_overview_sheet(wb, all_patients)
            
            # Save workbook
            wb.save(output_path)
            return True
            
        except Exception as e:
            raise Exception(f"Failed to generate all patients Excel report: {str(e)}")
            
    def _create_all_patients_summary_sheet(self, workbook, all_patients):
        """Create summary sheet for all patients with color-coded test results"""
        ws = workbook.create_sheet("All Patients Summary", 0)
        
        # Title
        ws['A1'] = "MEDICAL TEST RESULTS - ALL PATIENTS"
        ws['A1'].font = Font(bold=True, size=16)
        ws.merge_cells('A1:L1')
        
        # Headers
        headers = ['Patient ID', 'Name', 'Age', 'Gender', 'Test Name', 'Test Value', 
                  'Unit', 'Normal Range', 'Status', 'Test Date', 'Lab Technician', 'Notes']
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = self.fonts['header']
            cell.fill = self.colors['header']
            
        # Add border to headers
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'), 
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        current_row = 4
        
        # Process each patient
        for patient in all_patients:
            patient_id, first_name, last_name, dob, gender, phone, email, address, created_date = patient
            
            # Calculate age
            age = self.db_manager.calculate_age(dob) if dob else None
            age_display = f"{age}" if age else "N/A"
            
            # Get patient's test results
            test_results = self.db_manager.get_patient_test_results(patient_id)
            
            if not test_results:
                # Patient with no test results
                patient_name = f"{first_name or 'Unknown'} {last_name or 'Unknown'}"
                ws.cell(row=current_row, column=1, value=patient_id)
                ws.cell(row=current_row, column=2, value=patient_name)
                ws.cell(row=current_row, column=3, value=age_display)
                ws.cell(row=current_row, column=4, value=gender or 'N/A')
                ws.cell(row=current_row, column=5, value="No test results")
                current_row += 1
            else:
                # Patient with test results
                for result in test_results:
                    result_id, _, test_name, test_value, normal_min, normal_max, unit, test_date, lab_tech, notes = result
                    
                    # Get age/gender adjusted ranges
                    range_info = self.db_manager.get_age_gender_adjusted_range(test_name, age, gender)
                    if range_info['normal_min'] is not None and range_info['normal_max'] is not None:
                        normal_min, normal_max = range_info['normal_min'], range_info['normal_max']
                    
                    # Determine status and color
                    status = self.determine_test_status(test_value, normal_min, normal_max, test_name, age, gender)
                    color = self._get_status_color(status)
                    
                    # Format normal range
                    if normal_min is not None and normal_max is not None:
                        normal_range = f"{normal_min:.1f} - {normal_max:.1f}"
                    else:
                        normal_range = "Not defined"
                    
                    # Add data to row
                    patient_name = f"{first_name or 'Unknown'} {last_name or 'Unknown'}"
                    row_data = [
                        patient_id,
                        patient_name,
                        age_display,
                        gender or 'N/A',
                        test_name,
                        f"{test_value:.2f}",
                        unit or '',
                        normal_range,
                        status.replace('_', ' ').title(),
                        test_date,
                        lab_tech or '',
                        notes or ''
                    ]
                    
                    for col, value in enumerate(row_data, 1):
                        cell = ws.cell(row=current_row, column=col, value=value)
                        cell.border = thin_border
                        
                        # Apply color coding to test value and status columns
                        if col in [6, 9]:  # Test Value and Status columns
                            cell.fill = color
                            if color == self.colors['critical_high'] or color == self.colors['critical_low']:
                                cell.font = Font(color='FFFFFF', bold=True)  # White text for dark backgrounds
                    
                    current_row += 1
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = None
            for cell in column:
                try:
                    # Skip merged cells
                    if hasattr(cell, 'column_letter'):
                        if column_letter is None:
                            column_letter = cell.column_letter
                        if cell.value and len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                except:
                    pass
            
            if column_letter:
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
    def _create_patient_detail_sheet(self, workbook, patient, test_results):
        """Create detailed sheet for individual patient"""
        patient_id, first_name, last_name, dob, gender, phone, email, address, created_date = patient
        sheet_name = f"{patient_id}_{(first_name or 'Unknown')[:8]}"[:31]  # Excel sheet name limit
        
        ws = workbook.create_sheet(sheet_name)
        
        # Patient info header
        ws['A1'] = f"PATIENT DETAILS: {first_name or 'Unknown'} {last_name or 'Unknown'} (ID: {patient_id})"
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:H1')
        
        # Calculate age
        age = self.db_manager.calculate_age(dob) if dob else None
        
        # Patient information
        ws['A3'] = "Patient Information:"
        ws['A3'].font = self.fonts['header']
        
        info_data = [
            f"Date of Birth: {dob or 'Not provided'}",
            f"Age: {age if age else 'Unknown'} years",
            f"Gender: {gender or 'Not specified'}",
            f"Phone: {phone or 'Not provided'}",
            f"Email: {email or 'Not provided'}"
        ]
        
        for i, info in enumerate(info_data):
            ws[f'A{4+i}'] = info
        
        # Test results header
        ws['A10'] = "Test Results:"
        ws['A10'].font = self.fonts['header']
        
        # Test results table headers
        headers = ['Test Name', 'Value', 'Unit', 'Normal Range', 'Status', 'Date', 'Lab Tech', 'Notes']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=11, column=col, value=header)
            cell.font = self.fonts['header']
            cell.fill = self.colors['header']
        
        # Add test results with color coding
        for i, result in enumerate(test_results, 12):
            result_id, _, test_name, test_value, normal_min, normal_max, unit, test_date, lab_tech, notes = result
            
            # Get age/gender adjusted ranges
            range_info = self.db_manager.get_age_gender_adjusted_range(test_name, age, gender)
            if range_info['normal_min'] is not None and range_info['normal_max'] is not None:
                normal_min, normal_max = range_info['normal_min'], range_info['normal_max']
            
            # Determine status and color
            status = self.determine_test_status(test_value, normal_min, normal_max, test_name, age, gender)
            color = self._get_status_color(status)
            
            # Format normal range
            if normal_min is not None and normal_max is not None:
                normal_range = f"{normal_min:.1f} - {normal_max:.1f}"
            else:
                normal_range = "Not defined"
            
            # Add row data
            row_data = [
                test_name,
                f"{test_value:.2f}",
                unit or '',
                normal_range,
                status.replace('_', ' ').title(),
                test_date,
                lab_tech or '',
                notes or ''
            ]
            
            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=i, column=col, value=value)
                
                # Apply color coding to value and status columns
                if col in [2, 5]:  # Value and Status columns
                    cell.fill = color
                    if color == self.colors['critical_high'] or color == self.colors['critical_low']:
                        cell.font = Font(color='FFFFFF', bold=True)
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = None
            for cell in column:
                try:
                    # Skip merged cells
                    if hasattr(cell, 'column_letter'):
                        if column_letter is None:
                            column_letter = cell.column_letter
                        if cell.value and len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                except:
                    pass
            
            if column_letter:
                adjusted_width = min(max_length + 2, 30)
                ws.column_dimensions[column_letter].width = adjusted_width
            
    def _create_statistics_overview_sheet(self, workbook, all_patients):
        """Create statistics overview sheet"""
        ws = workbook.create_sheet("Statistics Overview")
        
        # Title
        ws['A1'] = "MEDICAL TEST STATISTICS OVERVIEW"
        ws['A1'].font = Font(bold=True, size=16)
        ws.merge_cells('A1:E1')
        
        # Basic statistics
        total_patients = len(all_patients)
        patients_with_tests = 0
        total_test_results = 0
        abnormal_results = 0
        
        # Count statistics
        for patient in all_patients:
            patient_id = patient[0]
            test_results = self.db_manager.get_patient_test_results(patient_id)
            if test_results:
                patients_with_tests += 1
                total_test_results += len(test_results)
                
                # Count abnormal results
                for result in test_results:
                    _, _, test_name, test_value, normal_min, normal_max, _, _, _, _ = result
                    age = self.db_manager.calculate_age(patient[3]) if patient[3] else None
                    gender = patient[4]
                    
                    status = self.determine_test_status(test_value, normal_min, normal_max, test_name, age, gender)
                    if status != 'normal':
                        abnormal_results += 1
        
        # Statistics display
        stats_data = [
            ("Total Patients:", total_patients),
            ("Patients with Test Results:", patients_with_tests),
            ("Total Test Results:", total_test_results),
            ("Normal Results:", total_test_results - abnormal_results),
            ("Abnormal Results:", abnormal_results),
            ("Abnormal Percentage:", f"{(abnormal_results/total_test_results*100):.1f}%" if total_test_results > 0 else "0%")
        ]
        
        ws['A3'] = "Overall Statistics:"
        ws['A3'].font = self.fonts['header']
        
        for i, (label, value) in enumerate(stats_data):
            ws[f'A{4+i}'] = label
            ws[f'B{4+i}'] = value
            ws[f'A{4+i}'].font = self.fonts['bold']
            
    def _get_status_color(self, status: str):
        """Get the appropriate color for a test status"""
        status_colors = {
            'normal': self.colors['normal'],
            'high': self.colors['high'],
            'low': self.colors['low'],
            'critical_high': self.colors['critical_high'],
            'critical_low': self.colors['critical_low'],
            'severely_high': self.colors['critical_high'],
            'severely_low': self.colors['critical_low']
        }
        return status_colors.get(status, self.colors['normal'])
    
    def _create_patient_summary_sheet(self, workbook, patient_info, test_results):
        """Create patient summary worksheet"""
        ws = workbook.create_sheet("Patient Summary")
        
        # Calculate patient age
        age = self.db_manager.calculate_age(patient_info[3]) if patient_info[3] else None
        gender = patient_info[4] if patient_info[4] else None
        
        # Patient Information Section
        ws['A1'] = "MEDICAL TEST REPORT"
        ws['A1'].font = Font(bold=True, size=16)
        ws.merge_cells('A1:E1')
        
        ws['A3'] = "Patient Information"
        ws['A3'].font = self.fonts['header']
        ws['A3'].fill = self.colors['header']
        
        # Patient details with age
        age_display = f"{age} years" if age is not None else "Unknown (no date of birth provided)"
        gender_display = gender if gender else "Not specified"
        
        patient_details = [
            ("Patient ID:", patient_info[0]),
            ("Name:", f"{patient_info[1] or 'Not provided'} {patient_info[2] or 'Not provided'}"),
            ("Date of Birth:", patient_info[3] if patient_info[3] else "Not provided"),
            ("Age:", age_display),
            ("Gender:", gender_display),
            ("Phone:", patient_info[5] if patient_info[5] else "Not provided"),
            ("Email:", patient_info[6] if patient_info[6] else "Not provided")
        ]
        
        for i, (label, value) in enumerate(patient_details):
            ws[f'A{4+i}'] = label
            ws[f'B{4+i}'] = value
            ws[f'A{4+i}'].font = self.fonts['bold']
        
        # Latest Test Results Summary with age/gender adjustments
        latest_results = self._get_latest_test_results(test_results)
        
        # Add note about age/gender adjustments
        adjustment_note = ""
        if age is None and gender is None:
            adjustment_note = " (Using general adult ranges - no age/gender info)"
        elif age is None:
            adjustment_note = f" (Using gender-adjusted ranges for {gender} - no age info)"
        elif gender is None:
            adjustment_note = f" (Using age-adjusted ranges for {age} years - no gender info)"
        else:
            adjustment_note = f" (Age/gender adjusted for {age}-year-old {gender})"
        
        ws['A13'] = f"Latest Test Results{adjustment_note}"
        ws['A13'].font = self.fonts['header']
        ws['A13'].fill = self.colors['header']
        
        # Headers for results table
        headers = ['Test Name', 'Value', 'Unit', 'Normal Range', 'Status', 'Range Source', 'Date']
        for i, header in enumerate(headers):
            cell = ws[f'{chr(65+i)}14']
            cell.value = header
            cell.font = self.fonts['header']
            cell.fill = self.colors['header']
        
        # Add latest results with age/gender adjustments
        for i, result in enumerate(latest_results):
            row = 15 + i
            test_name, test_value, normal_min, normal_max, unit, test_date = result
            
            # Get age/gender adjusted ranges
            range_info = self.db_manager.get_age_gender_adjusted_range(test_name, age, gender)
            if range_info['normal_min'] is not None and range_info['normal_max'] is not None:
                normal_min, normal_max = range_info['normal_min'], range_info['normal_max']
            
            ws[f'A{row}'] = test_name
            ws[f'B{row}'] = test_value
            ws[f'C{row}'] = unit if unit else ''
            
            if normal_min is not None and normal_max is not None:
                ws[f'D{row}'] = f"{normal_min} - {normal_max}"
                status = self.determine_test_status(test_value, normal_min, normal_max, test_name, age, gender)
                ws[f'E{row}'] = status.replace('_', ' ').title()
                
                # Apply color coding
                status_cell = ws[f'E{row}']
                if status in self.colors:
                    status_cell.fill = self.colors[status]
                
                # Add range source information
                ws[f'F{row}'] = range_info.get('source', 'Base range')
            else:
                ws[f'D{row}'] = "Not defined"
                ws[f'E{row}'] = "Unknown"
                ws[f'F{row}'] = "No range available"
            
            ws[f'G{row}'] = test_date
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _create_detailed_results_sheet(self, workbook, test_results):
        """Create detailed results worksheet"""
        ws = workbook.create_sheet("Detailed Results")
        
        # Get patient info for age/gender calculations
        if test_results:
            patient_id = test_results[0][1]  # Get patient_id from first result
            patient_info = self.db_manager.get_patient(patient_id)
            age = self.db_manager.calculate_age(patient_info[3]) if patient_info and patient_info[3] else None
            gender = patient_info[4] if patient_info and patient_info[4] else None
        else:
            age, gender = None, None
        
        # Convert test results to DataFrame for easier manipulation
        df_data = []
        for result in test_results:
            result_id, patient_id, test_name, test_value, normal_min, normal_max, unit, test_date, lab_tech, notes = result
            
            # Get age/gender adjusted ranges
            range_info = self.db_manager.get_age_gender_adjusted_range(test_name, age, gender)
            if range_info['normal_min'] is not None and range_info['normal_max'] is not None:
                normal_min, normal_max = range_info['normal_min'], range_info['normal_max']
            
            df_data.append({
                'Result ID': result_id,
                'Test Name': test_name,
                'Value': test_value,
                'Unit': unit if unit else '',
                'Normal Min': normal_min if normal_min is not None else '',
                'Normal Max': normal_max if normal_max is not None else '',
                'Test Date': test_date,
                'Lab Technician': lab_tech if lab_tech else '',
                'Notes': notes if notes else ''
            })
        
        df = pd.DataFrame(df_data)
        
        # Add headers
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)
        
        # Format headers
        for cell in ws[1]:
            cell.font = self.fonts['header']
            cell.fill = self.colors['header']
        
        # Apply color coding to test values with age/gender adjustments
        for row in range(2, len(df) + 2):
            test_value = ws[f'C{row}'].value
            normal_min = ws[f'E{row}'].value
            normal_max = ws[f'F{row}'].value
            test_name = ws[f'B{row}'].value
            
            if normal_min and normal_max and isinstance(test_value, (int, float)):
                try:
                    status = self.determine_test_status(float(test_value), float(normal_min), float(normal_max), 
                                                      test_name, age, gender)
                    if status in self.colors:
                        ws[f'C{row}'].fill = self.colors[status]
                except:
                    pass
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _create_trends_sheet(self, workbook, test_results):
        """Create trends analysis worksheet with charts"""
        ws = workbook.create_sheet("Trends Analysis")
        
        # Group results by test type
        test_groups = {}
        for result in test_results:
            test_name = result[2]
            if test_name not in test_groups:
                test_groups[test_name] = []
            test_groups[test_name].append(result)
        
        # Create trend data for tests with multiple results
        trend_data = []
        for test_name, results in test_groups.items():
            if len(results) > 1:
                # Sort by date
                sorted_results = sorted(results, key=lambda x: x[7])  # Sort by test_date
                for result in sorted_results:
                    trend_data.append({
                        'Test Name': test_name,
                        'Date': result[7],
                        'Value': result[3],
                        'Normal Min': result[4] if result[4] is not None else '',
                        'Normal Max': result[5] if result[5] is not None else ''
                    })
        
        if trend_data:
            df = pd.DataFrame(trend_data)
            
            # Add data to worksheet
            for r in dataframe_to_rows(df, index=False, header=True):
                ws.append(r)
            
            # Format headers
            for cell in ws[1]:
                cell.font = self.fonts['header']
                cell.fill = self.colors['header']
        else:
            ws['A1'] = "No trend data available (requires multiple test results of the same type)"
            ws['A1'].font = self.fonts['header']
    
    def _get_latest_test_results(self, test_results):
        """Get the latest result for each test type"""
        latest_results = {}
        
        for result in test_results:
            test_name = result[2]
            test_date = result[7]
            
            if test_name not in latest_results or test_date > latest_results[test_name][5]:
                latest_results[test_name] = (
                    test_name,           # 0: test_name
                    result[3],           # 1: test_value
                    result[4],           # 2: normal_min
                    result[5],           # 3: normal_max
                    result[6],           # 4: unit
                    test_date            # 5: test_date
                )
        
        return list(latest_results.values())
    
    def generate_patient_summary(self, patient_id: str, output_path: str) -> bool:
        """Generate a prescription-style patient test report (PDF)"""
        try:
            import matplotlib.pyplot as plt
            from matplotlib.backends.backend_pdf import PdfPages
            import matplotlib.patches as patches
            
            # Get patient and test data
            patient_info = self.db_manager.get_patient(patient_id)
            if not patient_info:
                raise ValueError(f"Patient {patient_id} not found")
            
            test_results = self.db_manager.get_patient_test_results_with_method(patient_id)
            if not test_results:
                raise ValueError(f"No test results found for patient {patient_id}")
            
            with PdfPages(output_path) as pdf:
                # Create prescription-style report
                fig, ax = plt.subplots(1, 1, figsize=(8.5, 11))
                ax.axis('off')
                
                # Current Y position for text placement
                y_pos = 0.95
                
                # 1. HEADER SECTION - Bold title with address
                ax.text(0.5, y_pos, 'RANDOMTEXT HOSPITAL', 
                       fontsize=18, fontweight='bold', ha='center', transform=ax.transAxes)
                y_pos -= 0.04
                ax.text(0.5, y_pos, 'RANDOMTEXT ADDRESS', 
                       fontsize=12, ha='center', transform=ax.transAxes)
                y_pos -= 0.04
                
                # 2. DEPARTMENT SECTION
                ax.text(0.05, y_pos, 'Department XYZ', 
                       fontsize=12, fontweight='bold', transform=ax.transAxes)
                y_pos -= 0.03
                
                # 3. PATIENT INFORMATION BOX
                # Create a box around patient info
                box_height = 0.12
                box_width = 0.9
                box_x = 0.05
                box_y = y_pos - box_height
                
                # Draw box border
                rect = patches.Rectangle((box_x, box_y), box_width, box_height, 
                                       linewidth=1, edgecolor='black', facecolor='none',
                                       transform=ax.transAxes)
                ax.add_patch(rect)
                
                # Calculate age
                age = self.db_manager.calculate_age(patient_info[3]) if patient_info[3] else "N/A"
                
                # Patient info inside the box
                info_y = y_pos - 0.02
                ax.text(0.07, info_y, f'Patient Name: {patient_info[1]} {patient_info[2]}', 
                       fontsize=11, fontweight='bold', transform=ax.transAxes)
                info_y -= 0.03
                ax.text(0.07, info_y, f'Date: {datetime.now().strftime("%d/%m/%Y")}', 
                       fontsize=10, transform=ax.transAxes)
                ax.text(0.5, info_y, f'Doctor: RANDOMTEXT', 
                       fontsize=10, transform=ax.transAxes)
                info_y -= 0.03
                ax.text(0.07, info_y, f'Age/Sex: {age}/{patient_info[4] if patient_info[4] else "N/A"}', 
                       fontsize=10, transform=ax.transAxes)
                ax.text(0.5, info_y, f'Ward: RANDOMTEXT', 
                       fontsize=10, transform=ax.transAxes)
                
                y_pos = box_y - 0.04
                
                # 4. TEST RESULTS TABLE (4 columns now)
                # Table headers with adjusted positions for 4 columns
                header_y = y_pos
                col_positions = [0.05, 0.28, 0.53, 0.78]
                headers = ['Parameter', 'Result Value', 'Normal Value', 'Method']
                
                for i, header in enumerate(headers):
                    ax.text(col_positions[i], header_y, header, fontsize=10, fontweight='bold', transform=ax.transAxes)
                
                # Draw line under headers
                ax.plot([0.05, 0.95], [header_y - 0.01, header_y - 0.01], 'k-', linewidth=0.5, transform=ax.transAxes)
                
                y_pos -= 0.03
                
                # Get latest test results and display only tests the patient has taken
                latest_results = self._get_latest_test_results_with_method(test_results)
                
                for result in latest_results:
                    test_name, test_value, normal_min, normal_max, unit, test_date, method = result
                    
                    # Get age/gender adjusted ranges
                    age_num = self.db_manager.calculate_age(patient_info[3]) if patient_info[3] else None
                    gender = patient_info[4] if patient_info[4] else None
                    range_info = self.db_manager.get_age_gender_adjusted_range(test_name, age_num, gender)
                    
                    if range_info['normal_min'] is not None and range_info['normal_max'] is not None:
                        normal_min, normal_max = range_info['normal_min'], range_info['normal_max']
                    
                    # Determine if result is abnormal
                    is_abnormal = False
                    if normal_min is not None and normal_max is not None:
                        status = self.determine_test_status(test_value, normal_min, normal_max, test_name, age_num, gender)
                        is_abnormal = status != 'normal'
                    
                    # Format values with units
                    unit_text = f" {unit}" if unit else ""
                    result_text = f"{test_value:.1f}{unit_text}"
                    
                    # Format normal range
                    if normal_min is not None and normal_max is not None:
                        normal_text = f"{normal_min:.1f}-{normal_max:.1f}{unit_text}"
                    else:
                        normal_text = "N/A"
                    
                    # Method text
                    method_text = method if method else "Standard Method"
                    
                    # Display row data
                    row_data = [test_name, result_text, normal_text, method_text]
                    
                    for i, data in enumerate(row_data):
                        # Special formatting for abnormal result values (column 1)
                        if i == 1 and is_abnormal:
                            # For abnormal values: bold, italic, underlined
                            ax.text(col_positions[i], y_pos, data, fontsize=9, fontweight='bold', 
                                   fontstyle='italic', transform=ax.transAxes)
                            # Add underline manually
                            text_width = len(data) * 0.006
                            ax.plot([col_positions[i], col_positions[i] + text_width], 
                                   [y_pos - 0.005, y_pos - 0.005], 'k-', linewidth=0.5, transform=ax.transAxes)
                        else:
                            ax.text(col_positions[i], y_pos, data, fontsize=9, transform=ax.transAxes)
                    
                    y_pos -= 0.025
                
                # 5. FOOTER SECTION
                # Move to bottom of page for footer
                footer_y = 0.15
                
                # Ending part placeholder
                ax.text(0.05, footer_y, 'ENDING PART XYZQWERTY', 
                       fontsize=10, transform=ax.transAxes)
                footer_y -= 0.04
                
                # End of report marker
                ax.text(0.5, footer_y, '****END OF REPORT****', 
                       fontsize=12, fontweight='bold', ha='center', transform=ax.transAxes)
                
                plt.tight_layout()
                pdf.savefig(fig, bbox_inches='tight')
                plt.close()
            
            return True
            
        except Exception as e:
            raise Exception(f"Failed to generate patient summary: {str(e)}")
    
    def _create_summary_chart(self, ax, test_results):
        """Create a summary chart for the PDF report"""
        ax.set_title('Test Results Timeline', fontweight='bold')
        
        # Group by test type and get recent results
        test_groups = {}
        for result in test_results:
            test_name = result[2]
            if test_name not in test_groups:
                test_groups[test_name] = []
            test_groups[test_name].append((result[7], result[3]))  # date, value
        
        colors = plt.cm.Set3(np.linspace(0, 1, len(test_groups)))
        
        for i, (test_name, results) in enumerate(list(test_groups.items())[:5]):  # Show top 5 tests
            if len(results) > 1:
                dates = [datetime.strptime(r[0], '%Y-%m-%d') for r in sorted(results)]
                values = [r[1] for r in sorted(results)]
                ax.plot(dates, values, marker='o', label=test_name[:15], color=colors[i])
        
        ax.legend(bbox_to_anchor=(1.05, 1), loc='upper left', fontsize=8)
        ax.set_xlabel('Date')
        ax.set_ylabel('Test Values')
        ax.grid(True, alpha=0.3)
        
        # Format x-axis dates
        if len(test_groups) > 0:
            ax.tick_params(axis='x', rotation=45)
    
    def _create_abnormal_results_summary(self, ax, latest_results, patient_info=None):
        """Create abnormal results summary for PDF"""
        ax.set_title('Abnormal Results Alert (Age/Gender Adjusted)', fontweight='bold')
        ax.axis('off')
        
        # Get age and gender for adjustments
        age = None
        gender = None
        if patient_info:
            age = self.db_manager.calculate_age(patient_info[3]) if patient_info[3] else None
            gender = patient_info[4] if patient_info[4] else None
        
        abnormal_count = {'high': 0, 'low': 0, 'critical_high': 0, 'critical_low': 0}
        abnormal_details = []
        
        for result in latest_results:
            test_name, test_value, normal_min, normal_max, unit, test_date = result
            
            # Get age/gender adjusted ranges
            range_info = self.db_manager.get_age_gender_adjusted_range(test_name, age, gender)
            if range_info['normal_min'] is not None and range_info['normal_max'] is not None:
                normal_min, normal_max = range_info['normal_min'], range_info['normal_max']
            
            if normal_min is not None and normal_max is not None:
                status = self.determine_test_status(test_value, normal_min, normal_max, test_name, age, gender)
                if status != 'normal':
                    abnormal_count[status] += 1
                    abnormal_details.append(f"{test_name}: {test_value} {unit or ''} ({status.replace('_', ' ').title()})")
        
        if any(abnormal_count.values()):
            alert_text = "⚠️ ABNORMAL RESULTS DETECTED ⚠️\n\n"
            alert_text += f"Critical High: {abnormal_count['critical_high']}\n"
            alert_text += f"High: {abnormal_count['high']}\n"
            alert_text += f"Low: {abnormal_count['low']}\n"
            alert_text += f"Critical Low: {abnormal_count['critical_low']}\n\n"
            alert_text += "Details (Age/Gender Adjusted):\n" + "\n".join(abnormal_details[:5])
        else:
            alert_text = "✅ All test results within normal ranges\n(Age/Gender Adjusted)"
        
        ax.text(0.1, 0.9, alert_text, transform=ax.transAxes, fontsize=9, 
                verticalalignment='top', fontfamily='monospace')
    
    def generate_test_statistics_report(self, output_path: str) -> bool:
        """Generate a system-wide statistics report"""
        try:
            # Get database statistics
            stats = self.db_manager.get_database_stats()
            
            # Create workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "System Statistics"
            
            # Add title
            ws['A1'] = "Medical Test System Statistics"
            ws['A1'].font = Font(bold=True, size=16)
            ws.merge_cells('A1:C1')
            
            # Add statistics
            stats_data = [
                ("Total Patients:", stats['patients']),
                ("Total Test Results:", stats['test_results']),
                ("Available Test Types:", stats['test_types']),
                ("Report Generated:", datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
            ]
            
            for i, (label, value) in enumerate(stats_data):
                ws[f'A{3+i}'] = label
                ws[f'B{3+i}'] = value
                ws[f'A{3+i}'].font = self.fonts['bold']
            
            wb.save(output_path)
            return True
            
        except Exception as e:
            raise Exception(f"Failed to generate statistics report: {str(e)}")
    
    def _get_latest_test_results_with_method(self, test_results):
        """
        Convert test results to include method information.
        Expected format: (test_name, test_value, normal_min, normal_max, unit, test_date, method)
        """
        formatted_results = []
        
        for result in test_results:
            # Extract data from the result tuple
            # Format: (test_result_id, patient_id, test_name, test_value, normal_min, normal_max, unit, test_date, lab_tech, notes, method)
            if len(result) >= 11:  # New format with method
                test_name = result[2]
                test_value = result[3]
                normal_min = result[4]
                normal_max = result[5]
                unit = result[6]
                test_date = result[7]
                method = result[10]
            else:  # Old format without method
                test_name = result[2]
                test_value = result[3]
                normal_min = result[4]
                normal_max = result[5]
                unit = result[6]
                test_date = result[7]
                method = "Standard Method"  # Default method
            
            formatted_results.append((test_name, test_value, normal_min, normal_max, unit, test_date, method))
        
        return formatted_results
