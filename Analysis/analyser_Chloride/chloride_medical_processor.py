"""
Simple Chloride Medical Data Processor
Focus: Only Chloride analysis from EIPL CSV files
"""

import pandas as pd
import numpy as np
import os
import logging
from datetime import datetime
import re
from pathlib import Path

# Try to import openpyxl for Excel functionality
try:
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

# Setup logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

# Chloride Clinical Ranges (Standard Medical Guidelines)
CHLORIDE_CLINICAL_RANGES = {
    'units': 'mEq/L (mmol/L)',
    'Normal': (96, 106),
    'Mild_Hypochloremia': (90, 95),
    'Severe_Hypochloremia': (0, 89),
    'Mild_Hyperchloremia': (107, 115),
    'Severe_Hyperchloremia': (116, 999)
}

# Risk Level Colors
RISK_COLORS = {
    'Normal': {'color': 'GREEN', 'symbol': '[OK]'},
    'Mild_Hypochloremia': {'color': 'YELLOW', 'symbol': '[CAUTION]'},
    'Severe_Hypochloremia': {'color': 'RED', 'symbol': '[URGENT]'},
    'Mild_Hyperchloremia': {'color': 'ORANGE', 'symbol': '[WATCH]'},
    'Severe_Hyperchloremia': {'color': 'PURPLE', 'symbol': '[CRITICAL]'}
}

# Excel Colors for actual cell backgrounds
EXCEL_COLORS = {
    'Normal': {'fill': '90EE90', 'font': '006400'},  # Light green background, dark green text
    'Mild_Hypochloremia': {'fill': 'FFFF99', 'font': 'B8860B'},  # Light yellow background, dark goldenrod text
    'Severe_Hypochloremia': {'fill': 'FFB6C1', 'font': '8B0000'},  # Light red background, dark red text
    'Mild_Hyperchloremia': {'fill': 'FFE4B5', 'font': 'FF8C00'},  # Light orange background, dark orange text
    'Severe_Hyperchloremia': {'fill': 'DDA0DD', 'font': '4B0082'}  # Light purple background, indigo text
}

# Clinical Actions
CHLORIDE_CLINICAL_ACTIONS = {
    'Normal': 'Normal electrolyte balance, maintain hydration',
    'Mild_Hypochloremia': 'Monitor fluid intake, check kidney function',
    'Severe_Hypochloremia': 'URGENT: Evaluate for dehydration, kidney/heart issues',
    'Mild_Hyperchloremia': 'Monitor blood pressure, check kidney function',
    'Severe_Hyperchloremia': 'CRITICAL: Check for severe dehydration, kidney failure'
}

class ChlorideProcessor:
    """Simple Chloride Data Processor for EIPL CSV files"""
    
    def __init__(self, append_mode=False):
        """Initialize the processor"""
        self.eipl_csv_file = self.find_eipl_csv()
        self.output_raw = "chloride_raw_data.csv"
        self.output_analyzed = "chloride_analyzed_results.csv"
        self.append_mode = append_mode
        
        logger.info(f"[INIT] Chloride Processor initialized (append_mode={append_mode})")
    
    def find_eipl_csv(self):
        """Find EIPL CSV file in current directory"""
        current_dir = Path(".")
        
        # Look for EIPL files
        for file in current_dir.glob("*.csv"):
            if file.name.upper().startswith("EIPL") and "CHLORIDE" in file.name.upper():
                logger.info(f"[DETECT] Found EIPL Chloride CSV: {file}")
                return str(file)
        
        # If no Chloride-specific file, look for any EIPL file
        for file in current_dir.glob("*.csv"):
            if file.name.upper().startswith("EIPL"):
                logger.info(f"[DETECT] Found EIPL CSV: {file}")
                return str(file)
        
        logger.warning("[WARNING] No EIPL CSV file found")
        return None
    
    def load_existing_processed_data(self):
        """Load existing processed data to check for duplicates"""
        existing_raw = pd.DataFrame()
        existing_analyzed = pd.DataFrame()
        
        try:
            if os.path.exists(self.output_raw):
                existing_raw = pd.read_csv(self.output_raw)
                logger.info(f"[LOAD] Found existing raw data: {len(existing_raw)} records")
            
            if os.path.exists(self.output_analyzed):
                existing_analyzed = pd.read_csv(self.output_analyzed)
                logger.info(f"[LOAD] Found existing analyzed data: {len(existing_analyzed)} records")
                
        except Exception as e:
            logger.warning(f"[WARNING] Error loading existing data: {e}")
            
        return existing_raw, existing_analyzed
    
    def get_processed_patient_ids(self, existing_raw_data):
        """Get set of already processed patient IDs"""
        if existing_raw_data.empty:
            return set()
        
        processed_ids = set()
        if 'ID' in existing_raw_data.columns:
            processed_ids = set(existing_raw_data['ID'].astype(str))
        
        logger.info(f"[DUPLICATE_CHECK] Found {len(processed_ids)} already processed patient IDs")
        return processed_ids

    def process_eipl_csv(self):
        """Process EIPL CSV file and extract chloride data"""
        if not self.eipl_csv_file or not os.path.exists(self.eipl_csv_file):
            logger.error("[ERROR] EIPL CSV file not found")
            return pd.DataFrame()
        
        logger.info(f"[PROCESS] Processing {self.eipl_csv_file}")
        
        # Load existing data if in append mode
        existing_processed_ids = set()
        if self.append_mode:
            existing_raw, _ = self.load_existing_processed_data()
            existing_processed_ids = self.get_processed_patient_ids(existing_raw)
        
        try:
            # Read CSV with different encodings
            df = None
            for encoding in ['utf-8', 'latin-1', 'cp1252']:
                try:
                    df = pd.read_csv(self.eipl_csv_file, encoding=encoding)
                    logger.info(f"[LOAD] Loaded {len(df)} records with {encoding} encoding")
                    break
                except UnicodeDecodeError:
                    continue
            
            if df is None:
                logger.error("[ERROR] Could not read CSV file with any encoding")
                return pd.DataFrame()
            
            # Process records
            processed_records = []
            skipped_count = 0
            
            for idx, row in df.iterrows():
                try:
                    # Skip header rows and invalid rows
                    sr_no = row.get('Sr. No.')
                    if pd.isna(sr_no) or str(sr_no).strip() == '' or str(sr_no).strip() == 'Sr. No.':
                        continue
                    
                    # Check if it's a valid numeric Sr. No.
                    try:
                        sr_no_num = int(float(str(sr_no)))
                        if sr_no_num <= 0:
                            continue
                    except (ValueError, TypeError):
                        continue
                    
                    # Extract patient data
                    patient_id = str(row.get('Patient ID', f'EIPL_{idx}')).strip()
                    if patient_id == 'nan' or patient_id == '':
                        patient_id = f'EIPL_{idx}'
                    
                    # Skip if already processed (in append mode)
                    if self.append_mode and patient_id in existing_processed_ids:
                        logger.debug(f"[SKIP] Patient {patient_id} already processed")
                        skipped_count += 1
                        continue
                    
                    # Get chloride reading from BIO-CHEQ column
                    chloride_reading = None
                    
                    # Try different column names for chloride reading
                    if 'BIO-CHEQ ' in row and pd.notna(row['BIO-CHEQ ']):
                        chloride_reading = row['BIO-CHEQ ']
                    elif 'BIO-CHEQ' in row and pd.notna(row['BIO-CHEQ']):
                        chloride_reading = row['BIO-CHEQ']
                    elif 'Device value (in conc.)' in row and pd.notna(row['Device value (in conc.)']):
                        chloride_reading = row['Device value (in conc.)']
                    
                    # Skip if no valid reading
                    if chloride_reading is None or pd.isna(chloride_reading) or str(chloride_reading).strip() == '':
                        continue
                    
                    # Clean up reading
                    chloride_reading_str = str(chloride_reading).strip()
                    
                    # Extract numeric value
                    try:
                        chloride_value = float(chloride_reading_str)
                    except ValueError:
                        # Try to extract numeric part
                        numeric_match = re.search(r'(\d+\.?\d*)', chloride_reading_str)
                        if not numeric_match:
                            continue
                        chloride_value = float(numeric_match.group(1))
                    
                    # Add units if not present
                    if not any(unit in chloride_reading_str.lower() for unit in ['meq/l', 'mmol/l']):
                        chloride_reading_str = f"{chloride_value} mEq/L"
                    
                    # Get demographics
                    age = row.get('Age', 35)
                    try:
                        age = int(float(age)) if pd.notna(age) and str(age).upper() != 'NA' else 35
                    except (ValueError, TypeError):
                        age = 35
                    
                    gender = str(row.get('Gender', 'Male')).strip()
                    if gender.lower() in ['unknown', 'nan', '', 'na']:
                        gender = 'Male'  # Default
                    
                    # Get test type
                    test_type = str(row.get('Test type', 'Serum')).strip()
                    
                    # Create record
                    record = {
                        'ID': patient_id,
                        'Test_name': 'Chloride',
                        'reading': chloride_reading_str,
                        'numeric_value': chloride_value,
                        'Age': age,
                        'Gender': gender,
                        'Test_Type': test_type,
                        'Source_File': 'EIPL_BIO-CHEQ',
                        'Timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                    }
                    
                    processed_records.append(record)
                    
                except Exception as e:
                    logger.warning(f"[WARNING] Error processing row {idx}: {e}")
                    continue
            
            if not processed_records:
                if self.append_mode and skipped_count > 0:
                    logger.info(f"[APPEND] No new records found - {skipped_count} patients already processed")
                    print(f"[INFO] No new chloride data found - {skipped_count} patients already in Excel file")
                else:
                    logger.warning("[WARNING] No valid chloride records found")
                return pd.DataFrame()
            
            result_df = pd.DataFrame(processed_records)
            if self.append_mode:
                logger.info(f"[APPEND] Found {len(result_df)} NEW chloride records (skipped {skipped_count} existing)")
                print(f"[APPEND] Found {len(result_df)} new patients (skipped {skipped_count} existing)")
            else:
                logger.info(f"[SUCCESS] Processed {len(result_df)} chloride records")
            
            return result_df
            
        except Exception as e:
            logger.error(f"[ERROR] Error processing EIPL CSV: {e}")
            return pd.DataFrame()
    
    def classify_chloride(self, chloride_value):
        """Classify chloride result based on clinical ranges"""
        
        ranges = CHLORIDE_CLINICAL_RANGES
        
        # Classification logic:
        if chloride_value < 90:
            return 'Severe_Hypochloremia', ranges['Severe_Hypochloremia']
        elif chloride_value <= 95:
            return 'Mild_Hypochloremia', ranges['Mild_Hypochloremia']
        elif chloride_value <= 106:
            return 'Normal', ranges['Normal']
        elif chloride_value <= 115:
            return 'Mild_Hyperchloremia', ranges['Mild_Hyperchloremia']
        else:
            return 'Severe_Hyperchloremia', ranges['Severe_Hyperchloremia']
    
    def analyze_chloride_data(self, raw_data):
        """Analyze chloride data and create classifications"""
        if raw_data.empty:
            logger.warning("[WARNING] No data to analyze")
            return pd.DataFrame()
        
        logger.info(f"[ANALYZE] Analyzing {len(raw_data)} chloride records")
        
        analyzed_records = []
        
        for _, row in raw_data.iterrows():
            try:
                chloride_value = row['numeric_value']
                age = row['Age']
                gender = row['Gender']
                test_type = row.get('Test_Type', 'Serum')
                
                # Classify
                classification, classification_range = self.classify_chloride(chloride_value)
                
                # Get the normal range
                normal_range = CHLORIDE_CLINICAL_RANGES['Normal']
                
                # Get risk info
                risk_info = RISK_COLORS[classification]
                clinical_action = CHLORIDE_CLINICAL_ACTIONS[classification]
                
                # Create analyzed record (simplified)
                analyzed_record = {
                    'Patient_ID': row['ID'],
                    'Test_Name': row['Test_name'],
                    'Age': age,
                    'Gender': gender,
                    'Chloride_Value': f"{chloride_value:.1f} mEq/L",
                    'Test_Type': test_type,
                    'Classification': classification.replace('_', ' '),  # Make it readable
                    'Normal_Range': f"{normal_range[0]}-{normal_range[1]} mEq/L",
                    'Color_Code': risk_info['color']
                }
                
                analyzed_records.append(analyzed_record)
                
            except Exception as e:
                logger.error(f"[ERROR] Error analyzing record {row['ID']}: {e}")
                continue
        
        analyzed_df = pd.DataFrame(analyzed_records)
        logger.info(f"[SUCCESS] Analyzed {len(analyzed_df)} records")
        
        return analyzed_df
    
    def display_results(self, analyzed_df):
        """Display chloride analysis results"""
        if analyzed_df.empty:
            print("[INFO] No results to display")
            return
        
        print("\n" + "="*80)
        print("[REPORT] CHLORIDE ANALYSIS RESULTS")
        print("="*80)
        
        # Sort by classification severity
        classification_order = {'Normal': 1, 'Mild Hypochloremia': 2, 'Severe Hypochloremia': 3, 
                               'Mild Hyperchloremia': 4, 'Severe Hyperchloremia': 5}
        analyzed_df['sort_order'] = analyzed_df['Classification'].map(classification_order).fillna(2)
        sorted_results = analyzed_df.sort_values('sort_order', ascending=False)
        
        # Display individual results
        print("\n[INDIVIDUAL RESULTS]")
        print("-" * 50)
        
        for _, row in sorted_results.iterrows():
            # Get the original classification key for risk colors
            class_key = row['Classification'].replace(' ', '_')
            if class_key in RISK_COLORS:
                print(f"\n{RISK_COLORS[class_key]['symbol']} {row['Color_Code']} - {row['Classification']}")
                print(f"   [ID] Patient: {row['Patient_ID']}")
                print(f"   [TEST] Test: {row['Test_Name']} ({row.get('Test_Type', 'Serum')})")
                print(f"   [CHLORIDE] Chloride: {row['Chloride_Value']}")
                print(f"   [NORMAL] Range: {row['Normal_Range']}")
                print(f"   [DEMO] Age: {row['Age']}, Gender: {row['Gender']}")
        
        # Summary statistics
        print("\n" + "="*80)
        print("[SUMMARY] CHLORIDE CLASSIFICATION SUMMARY")
        print("="*80)
        
        total = len(analyzed_df)
        classification_counts = analyzed_df['Classification'].value_counts()
        
        print(f"\nTotal Patients Analyzed: {total}")
        print("\nClassification Distribution:")
        
        for classification in ['Normal', 'Mild Hypochloremia', 'Severe Hypochloremia', 
                              'Mild Hyperchloremia', 'Severe Hyperchloremia']:
            count = classification_counts.get(classification, 0)
            percentage = (count / total * 100) if total > 0 else 0
            class_key = classification.replace(' ', '_')
            if class_key in RISK_COLORS:
                risk_info = RISK_COLORS[class_key]
                print(f"   {risk_info['symbol']} {risk_info['color']} {classification}: {count} patients ({percentage:.1f}%)")
        
        # Gender analysis
        print("\n[GENDER ANALYSIS]")
        gender_groups = analyzed_df.groupby('Gender')
        for gender, group_data in gender_groups:
            abnormal_count = len(group_data[group_data['Classification'] != 'Normal'])
            total_count = len(group_data)
            percentage = (abnormal_count / total_count * 100) if total_count > 0 else 0
            print(f"   {gender}: {abnormal_count}/{total_count} with chloride abnormalities ({percentage:.1f}%)")
        
        # Age group analysis
        age_groups = {
            'Pediatric (0-12)': analyzed_df[analyzed_df['Age'] <= 12],
            'Adolescent (13-18)': analyzed_df[(analyzed_df['Age'] > 12) & (analyzed_df['Age'] <= 18)],
            'Adult (19-65)': analyzed_df[(analyzed_df['Age'] > 18) & (analyzed_df['Age'] <= 65)],
            'Elderly (65+)': analyzed_df[analyzed_df['Age'] > 65]
        }
        
        print("\n[AGE GROUP ANALYSIS]")
        for group_name, group_data in age_groups.items():
            if not group_data.empty:
                abnormal_count = len(group_data[group_data['Classification'] != 'Normal'])
                total_count = len(group_data)
                percentage = (abnormal_count / total_count * 100) if total_count > 0 else 0
                print(f"   {group_name}: {abnormal_count}/{total_count} with chloride abnormalities ({percentage:.1f}%)")
        
        # Critical alerts
        critical_cases = analyzed_df[analyzed_df['Classification'].isin(['Severe Hypochloremia', 'Severe Hyperchloremia'])]
        if not critical_cases.empty:
            print(f"\n🚨 CRITICAL ALERT: {len(critical_cases)} patients require IMMEDIATE attention!")
            
            severe_hypo = analyzed_df[analyzed_df['Classification'] == 'Severe Hypochloremia']
            if not severe_hypo.empty:
                print(f"\n🔴 SEVERE HYPOCHLOREMIA ({len(severe_hypo)}):")
                for _, case in severe_hypo.iterrows():
                    print(f"   🔴 Patient {case['Patient_ID']}: {case['Chloride_Value']} (Normal: {case['Normal_Range']})")
            
            severe_hyper = analyzed_df[analyzed_df['Classification'] == 'Severe Hyperchloremia']
            if not severe_hyper.empty:
                print(f"\n🟣 SEVERE HYPERCHLOREMIA ({len(severe_hyper)}):")
                for _, case in severe_hyper.iterrows():
                    print(f"   🟣 Patient {case['Patient_ID']}: {case['Chloride_Value']} (Normal: {case['Normal_Range']})")
    
    def save_excel_with_colors(self, analyzed_data):
        """Save analyzed data to Excel file with colored cells"""
        if not EXCEL_AVAILABLE:
            logger.warning("[WARNING] openpyxl not available - skipping Excel export")
            return False
        
        if analyzed_data.empty:
            logger.warning("[WARNING] No data to save to Excel")
            return False
        
        try:
            excel_file = "chloride_analyzed_results.xlsx"
            
            # Handle append mode - load existing Excel if it exists
            if self.append_mode and os.path.exists(excel_file):
                logger.info("[APPEND] Loading existing Excel file for append mode")
                try:
                    # Load existing Excel file
                    from openpyxl import load_workbook
                    wb = load_workbook(excel_file)
                    
                    # Get existing data worksheet
                    if "Chloride Analysis" in wb.sheetnames:
                        ws = wb["Chloride Analysis"]
                        
                        # Read existing data (skip the first 4 rows which are headers/title)
                        existing_data = []
                        for row in ws.iter_rows(min_row=5, values_only=True):
                            if row[0] is not None:  # If first column has data
                                existing_data.append(row)
                        
                        if existing_data:
                            # Convert to DataFrame to combine with new data
                            existing_df = pd.DataFrame(existing_data, columns=[
                                'Patient_ID', 'Test_Name', 'Age', 'Gender', 'Chloride_Value', 
                                'Classification', 'Normal_Range', 'Color_Code'
                            ])
                            
                            # Add gold standard columns if they exist
                            if len(existing_data[0]) > 8:
                                existing_df = pd.DataFrame(existing_data, columns=[
                                    'Patient_ID', 'Test_Name', 'Age', 'Gender', 'Chloride_Value', 
                                    'Classification', 'Normal_Range', 'Color_Code', 'Gold_Standard', 'Accuracy_Error'
                                ])
                            
                            # Combine existing and new data
                            combined_data = pd.concat([existing_df, analyzed_data], ignore_index=True)
                            # Remove duplicates based on Patient_ID (keep the new one)
                            combined_data = combined_data.drop_duplicates(subset=['Patient_ID'], keep='last')
                            analyzed_data = combined_data
                            logger.info(f"[APPEND] Combined {len(existing_df)} existing + {len(analyzed_data) - len(existing_df)} new records")
                    
                    # Remove the old worksheet and create fresh one with combined data
                    if "Chloride Analysis" in wb.sheetnames:
                        wb.remove(wb["Chloride Analysis"])
                    
                except Exception as e:
                    logger.warning(f"[WARNING] Could not load existing Excel for append: {e}")
                    logger.info("[FALLBACK] Creating new Excel file instead")
                    wb = Workbook()
            else:
                # Create new workbook
                wb = Workbook()
            
            # Create/recreate the main data worksheet
            ws = wb.create_sheet("Chloride Analysis")
            if "Sheet" in wb.sheetnames:
                wb.remove(wb["Sheet"])
            wb.active = ws
            
            # Remove the sort_order column if it exists for Excel export
            excel_data = analyzed_data.copy()
            if 'sort_order' in excel_data.columns:
                excel_data = excel_data.drop('sort_order', axis=1)
            
            # Write data to worksheet
            for r in dataframe_to_rows(excel_data, index=False, header=True):
                ws.append(r)
            
            # Style the header row
            header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            header_font = Font(color='FFFFFF', bold=True)
            header_alignment = Alignment(horizontal='center', vertical='center')
            
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment
            
            # Apply colors based on classification
            classification_col = None
            color_code_col = None
            
            # Find the column indices
            for idx, cell in enumerate(ws[1], 1):
                if cell.value == 'Classification':
                    classification_col = idx
                elif cell.value == 'Color_Code':
                    color_code_col = idx
            
            # Apply colors to data rows
            for row in range(2, ws.max_row + 1):
                if classification_col:
                    classification = ws.cell(row=row, column=classification_col).value
                    # Convert back to key format for color lookup
                    class_key = classification.replace(' ', '_') if classification else 'Normal'
                    if class_key in EXCEL_COLORS:
                        color_info = EXCEL_COLORS[class_key]
                        
                        # Color the entire row
                        for col in range(1, ws.max_column + 1):
                            cell = ws.cell(row=row, column=col)
                            cell.fill = PatternFill(start_color=color_info['fill'], 
                                                  end_color=color_info['fill'], 
                                                  fill_type='solid')
                            cell.font = Font(color=color_info['font'], bold=False)
                            cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Add borders
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            for row in ws.iter_rows():
                for cell in row:
                    cell.border = thin_border
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                
                adjusted_width = min(max_length + 2, 20)  # Cap at 20 characters
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Add a title and summary at the top
            ws.insert_rows(1, 3)
            
            # Title
            ws['A1'] = 'CHLORIDE ANALYSIS REPORT'
            ws['A1'].font = Font(size=16, bold=True, color='2F4F4F')
            ws['A1'].alignment = Alignment(horizontal='center')
            ws.merge_cells('A1:I1')
            
            # Timestamp
            ws['A2'] = f'Generated: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}'
            ws['A2'].font = Font(size=11, italic=True, color='696969')
            ws['A2'].alignment = Alignment(horizontal='center')
            ws.merge_cells('A2:I2')
            
            # Summary
            total_patients = len(analyzed_data)
            normal_count = len(analyzed_data[analyzed_data['Classification'] == 'Normal'])
            abnormal_count = len(analyzed_data[analyzed_data['Classification'] != 'Normal'])
            
            ws['A3'] = f'Total Patients: {total_patients} | Normal: {normal_count} | Abnormal Chloride: {abnormal_count}'
            ws['A3'].font = Font(size=11, bold=True, color='2F4F4F')
            ws['A3'].alignment = Alignment(horizontal='center')
            ws.merge_cells('A3:I3')
            
            # Protect the worksheet with password while allowing sort and filter
            ws.protection.password = "eipl"
            ws.protection.sheet = True
            ws.protection.sort = False  # Allow sorting
            ws.protection.autoFilter = False  # Allow filtering
            ws.protection.selectLockedCells = True  # Allow selecting cells
            ws.protection.selectUnlockedCells = True  # Allow selecting unlocked cells
            
            # Enable AutoFilter for the data range (starting from row 4 which has headers after our title rows)
            if ws.max_row > 4:  # Ensure we have data rows
                ws.auto_filter.ref = f"A4:{chr(64 + ws.max_column)}{ws.max_row}"
            
            # Create/Update Summary Report worksheet
            if "Summary Report" in wb.sheetnames:
                wb.remove(wb["Summary Report"])
            summary_ws = wb.create_sheet("Summary Report")
            self._create_summary_worksheet(summary_ws, analyzed_data)
            
            # Save the workbook
            wb.save(excel_file)
            
            if self.append_mode:
                logger.info(f"[SAVE] Excel file updated with new data (Protected with password)")
                print("   [APPEND] ✨ New data added to existing Excel file!")
            else:
                logger.info(f"[SAVE] Excel file with colors and summary report saved to {excel_file} (Protected with password)")
                print("   [SUCCESS] ✨ Excel file with colored data and summary report created!")
                
            return True
            
        except Exception as e:
            logger.error(f"[ERROR] Error creating Excel file: {e}")
            return False
    
    def _create_summary_worksheet(self, ws, analyzed_data):
        """Create a comprehensive summary report worksheet"""
        try:
            # Title
            ws['A1'] = 'CHLORIDE ANALYSIS SUMMARY REPORT'
            ws['A1'].font = Font(size=18, bold=True, color='2F4F4F')
            ws['A1'].alignment = Alignment(horizontal='center')
            ws.merge_cells('A1:D1')
            
            # Timestamp and mode
            ws['A3'] = f'Generated: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}'
            ws['A3'].font = Font(size=11, italic=True, color='696969')
            
            mode_text = "Mode: APPEND - Added new records" if self.append_mode else "Mode: COMPLETE ANALYSIS"
            ws['A4'] = mode_text
            ws['A4'].font = Font(size=11, color='696969')
            
            # Basic statistics
            total = len(analyzed_data)
            ws['A6'] = 'BASIC STATISTICS'
            ws['A6'].font = Font(size=14, bold=True, color='2F4F4F')
            
            ws['A7'] = f'Total Patients: {total}'
            ws['A7'].font = Font(size=11, bold=True)
            
            # Classification distribution
            ws['A9'] = 'CLASSIFICATION DISTRIBUTION'
            ws['A9'].font = Font(size=14, bold=True, color='2F4F4F')
            
            classification_counts = analyzed_data['Classification'].value_counts()
            row = 10
            # Use actual chloride classifications that the processor produces  
            chloride_classifications = ['Normal', 'Mild_Hypochloremia', 'Severe_Hypochloremia', 'Mild_Hyperchloremia', 'Severe_Hyperchloremia']
            
            for classification in chloride_classifications:
                count = classification_counts.get(classification, 0)
                percentage = (count / total * 100) if total > 0 else 0
                
                # Classification name
                display_name = classification.replace('_', ' ')
                ws[f'A{row}'] = f'{display_name}:'
                ws[f'A{row}'].font = Font(size=11, bold=True)
                
                # Count and percentage
                ws[f'B{row}'] = f'{count} patients ({percentage:.1f}%)'
                ws[f'B{row}'].font = Font(size=11)
                
                # Color coding based on severity
                if 'Normal' in classification:
                    ws[f'A{row}'].fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')
                    ws[f'B{row}'].fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')
                elif 'Severe' in classification:
                    ws[f'A{row}'].fill = PatternFill(start_color='FFB6C1', end_color='FFB6C1', fill_type='solid')
                    ws[f'B{row}'].fill = PatternFill(start_color='FFB6C1', end_color='FFB6C1', fill_type='solid')
                elif 'Mild' in classification:
                    ws[f'A{row}'].fill = PatternFill(start_color='FFFF99', end_color='FFFF99', fill_type='solid')
                    ws[f'B{row}'].fill = PatternFill(start_color='FFFF99', end_color='FFFF99', fill_type='solid')
                
                row += 1
            
            # Gender analysis
            row += 1
            ws[f'A{row}'] = 'GENDER ANALYSIS'
            ws[f'A{row}'].font = Font(size=14, bold=True, color='2F4F4F')
            row += 1
            
            gender_groups = analyzed_data.groupby('Gender')
            for gender, group_data in gender_groups:
                abnormal_count = len(group_data[group_data['Classification'] != 'Normal'])
                total_count = len(group_data)
                percentage = (abnormal_count / total_count * 100) if total_count > 0 else 0
                
                ws[f'A{row}'] = f'{gender}:'
                ws[f'A{row}'].font = Font(size=11, bold=True)
                ws[f'B{row}'] = f'{abnormal_count}/{total_count} with abnormal chloride ({percentage:.1f}%)'
                ws[f'B{row}'].font = Font(size=11)
                row += 1
            
            # Critical alerts
            severe_cases = analyzed_data[analyzed_data['Classification'].isin(['Severe_Hypochloremia', 'Severe_Hyperchloremia'])]
            if not severe_cases.empty:
                row += 1
                ws[f'A{row}'] = 'CRITICAL ALERTS'
                ws[f'A{row}'].font = Font(size=14, bold=True, color='8B0000')
                row += 1
                
                ws[f'A{row}'] = f'🚨 URGENT: {len(severe_cases)} patients with SEVERE chloride abnormalities!'
                ws[f'A{row}'].font = Font(size=12, bold=True, color='8B0000')
                ws[f'A{row}'].fill = PatternFill(start_color='FFB6C1', end_color='FFB6C1', fill_type='solid')
                row += 1
                
                for _, case in severe_cases.iterrows():
                    ws[f'A{row}'] = f'🔴 Patient {case["Patient_ID"]}:'
                    ws[f'A{row}'].font = Font(size=10, bold=True)
                    ws[f'B{row}'] = f'{case["Chloride_Value"]} (Normal: {case["Normal_Range"]})'
                    ws[f'B{row}'].font = Font(size=10)
                    ws[f'A{row}'].fill = PatternFill(start_color='FFB6C1', end_color='FFB6C1', fill_type='solid')
                    ws[f'B{row}'].fill = PatternFill(start_color='FFB6C1', end_color='FFB6C1', fill_type='solid')
                    row += 1
            
            # Auto-adjust column widths
            ws.column_dimensions['A'].width = 35
            ws.column_dimensions['B'].width = 30
            ws.column_dimensions['C'].width = 20
            ws.column_dimensions['D'].width = 20
            
            # Apply protection to summary worksheet
            ws.protection.password = "eipl"
            ws.protection.sheet = True
            ws.protection.selectLockedCells = True
            ws.protection.selectUnlockedCells = True
            
            logger.info("[SUMMARY] Chloride summary worksheet created successfully")
            
        except Exception as e:
            logger.error(f"[ERROR] Error creating chloride summary worksheet: {e}")
    
    def save_results(self, raw_data, analyzed_data):
        """Save results to Excel with colors and integrated summary (no separate files)"""
        try:
            if not analyzed_data.empty:
                # Only create Excel with colors and integrated summary
                excel_success = self.save_excel_with_colors(analyzed_data)
                if excel_success:
                    if not self.append_mode:
                        print("   [SUCCESS] ✨ Excel file with colored data and summary report created!")
                    print("   [INFO] 📊 Two worksheets: 'Chloride Analysis' (data) + 'Summary Report' (analysis)")
                    logger.info(f"[SAVE] Complete Excel analysis saved with {len(analyzed_data)} records and summary")
            
        except Exception as e:
            logger.error(f"[ERROR] Error saving results: {e}")
    
    def run_complete_analysis(self):
        """Run complete chloride analysis workflow"""
        print("\n" + "="*80)
        print("[SYSTEM] CHLORIDE MEDICAL DATA PROCESSOR")
        print("="*80)
        print(f"[START] Starting analysis at {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        
        try:
            # Step 1: Process EIPL CSV
            print("\n[STEP 1] Processing EIPL CSV file...")
            raw_data = self.process_eipl_csv()
            
            if raw_data.empty:
                print("[ERROR] No data found to process!")
                return False
            
            print(f"[SUCCESS] Processed {len(raw_data)} chloride records")
            
            # Step 2: Analyze data
            print("\n[STEP 2] Analyzing chloride data...")
            analyzed_data = self.analyze_chloride_data(raw_data)
            
            if analyzed_data.empty:
                print("[ERROR] Analysis failed!")
                return False
            
            print(f"[SUCCESS] Analyzed {len(analyzed_data)} records")
            
            # Step 3: Display results
            print("\n[STEP 3] Displaying results...")
            self.display_results(analyzed_data)
            
            # Step 4: Save results
            print("\n[STEP 4] Saving results...")
            self.save_results(raw_data, analyzed_data)
            
            print("\n" + "="*80)
            print("[COMPLETE] Chloride analysis completed successfully!")
            print("="*80)
            
            return True
            
        except Exception as e:
            logger.error(f"[ERROR] Analysis failed: {e}")
            print(f"\n[ERROR] Analysis failed: {e}")
            return False


def main():
    """Main function"""
    try:
        print("\n" + "="*80)
        print("[CHLORIDE PROCESSOR] Smart Excel Analysis")
        print("="*80)
        print("[INFO] This processor generates ONE Excel file with:")
        print("[INFO] • 'Chloride Analysis' sheet - Color-coded patient data")
        print("[INFO] • 'Summary Report' sheet - Complete analysis summary")
        print("[INFO] • Password protection with sorting/filtering enabled")
        print("="*80)
        print("[APPEND FEATURE] You can add new patients to your EIPL CSV")
        print("[APPEND FEATURE] and choose append mode to add them to existing Excel!")
        print("="*80)
        
        choice = input("\nChoose mode:\n[1] Complete analysis (overwrites existing Excel)\n[2] Append new data (adds to existing Excel)\nPress Enter for complete analysis: ").strip()
        
        append_mode = False
        if choice == "2":
            append_mode = True
            print("\n[MODE] APPEND MODE - Adding new patients to existing Excel file")
        else:
            append_mode = False
            print("\n[MODE] COMPLETE MODE - Generating new Excel file")
        
        processor = ChlorideProcessor(append_mode=append_mode)
        success = processor.run_complete_analysis()
        
        if success:
            print("\n[FILES] Generated file:")
            excel_file = "chloride_analyzed_results.xlsx"
            if os.path.exists(excel_file):
                file_size = os.path.getsize(excel_file)
                print(f"   [OK] {excel_file} ({file_size} bytes)")
                print("        ├── 'Chloride Analysis' sheet: Color-coded patient data with filtering")
                print("        └── 'Summary Report' sheet: Complete statistical analysis and recommendations")
        
    except Exception as e:
        print(f"[FATAL] Fatal error: {e}")
    
    print("\nPress Enter to exit...")
    input()


if __name__ == "__main__":
    main()
