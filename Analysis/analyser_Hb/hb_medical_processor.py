"""
Simple Hemoglobin Medical Data Processor
Focus: Only Hemoglobin analysis from EIPL CSV files
"""

import pandas as pd
import numpy as np
import os
import logging
from datetime import datetime
import re
from pathlib import Path

# Try to import openpyxl for Excel functionality
try:
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

# Setup logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

# Hemoglobin Clinical Ranges (Custom Standards)
HB_CLINICAL_RANGES = {
    'units': 'g/dL',
    'Male': {
        'Normal': (13.0, 17.0),
        'Mild Anemia': (10.0, 12.9),
        'Severe Anemia': (0, 9.9),
        'Hemoglobinemia': (17.1, 999.9)  # Above normal
    },
    'Female': {
        'Normal': (12.0, 15.0),
        'Mild Anemia': (10.0, 11.9),
        'Severe Anemia': (0, 9.9),
        'Hemoglobinemia': (15.1, 999.9)  # Above normal
    }
}

# Risk Level Colors
RISK_COLORS = {
    'Normal': {'color': 'GREEN', 'symbol': '[OK]'},
    'Mild Anemia': {'color': 'YELLOW', 'symbol': '[CAUTION]'},
    'Severe Anemia': {'color': 'RED', 'symbol': '[URGENT]'},
    'Hemoglobinemia': {'color': 'BLUE', 'symbol': '[HIGH]'}
}

# Excel Colors for actual cell backgrounds
EXCEL_COLORS = {
    'Normal': {'fill': '90EE90', 'font': '006400'},  # Light green background, dark green text
    'Mild Anemia': {'fill': 'FFFF99', 'font': 'B8860B'},  # Light yellow background, dark goldenrod text
    'Severe Anemia': {'fill': 'FFB6C1', 'font': '8B0000'},  # Light red background, dark red text
    'Hemoglobinemia': {'fill': 'ADD8E6', 'font': '000080'}  # Light blue background, navy text
}

class HemoglobinProcessor:
    """Simple Hemoglobin Data Processor for EIPL CSV files"""
    
    def __init__(self, append_mode=False):
        """Initialize the processor"""
        self.eipl_csv_file = self.find_eipl_csv()
        self.output_raw = "hb_raw_data.csv"
        self.output_analyzed = "hb_analyzed_results.csv"
        self.append_mode = append_mode
        
        logger.info(f"[INIT] Hemoglobin Processor initialized (append_mode={append_mode})")
    
    def find_eipl_csv(self):
        """Find EIPL CSV file in current directory"""
        current_dir = Path(".")
        
        # Look for EIPL files
        for file in current_dir.glob("*.csv"):
            if file.name.upper().startswith("EIPL") and "HB" in file.name.upper():
                logger.info(f"[DETECT] Found EIPL Hb CSV: {file}")
                return str(file)
        
        # If no Hb-specific file, look for any EIPL file
        for file in current_dir.glob("*.csv"):
            if file.name.upper().startswith("EIPL"):
                logger.info(f"[DETECT] Found EIPL CSV: {file}")
                return str(file)
        
        logger.warning("[WARNING] No EIPL CSV file found")
        return None
    
    def load_existing_processed_data(self):
        """Load existing processed data to check for duplicates"""
        existing_raw = pd.DataFrame()
        existing_analyzed = pd.DataFrame()
        
        try:
            if os.path.exists(self.output_raw):
                existing_raw = pd.read_csv(self.output_raw)
                logger.info(f"[LOAD] Found existing raw data: {len(existing_raw)} records")
            
            if os.path.exists(self.output_analyzed):
                existing_analyzed = pd.read_csv(self.output_analyzed)
                logger.info(f"[LOAD] Found existing analyzed data: {len(existing_analyzed)} records")
                
        except Exception as e:
            logger.warning(f"[WARNING] Error loading existing data: {e}")
            
        return existing_raw, existing_analyzed
    
    def get_processed_patient_ids(self, existing_raw_data):
        """Get set of already processed patient IDs"""
        if existing_raw_data.empty:
            return set()
        
        processed_ids = set()
        if 'ID' in existing_raw_data.columns:
            processed_ids = set(existing_raw_data['ID'].astype(str))
        
        logger.info(f"[DUPLICATE_CHECK] Found {len(processed_ids)} already processed patient IDs")
        return processed_ids
    
    def process_eipl_csv(self):
        """Process EIPL CSV file and extract hemoglobin data"""
        if not self.eipl_csv_file or not os.path.exists(self.eipl_csv_file):
            logger.error("[ERROR] EIPL CSV file not found")
            return pd.DataFrame()
        
        logger.info(f"[PROCESS] Processing {self.eipl_csv_file}")
        
        # Load existing data if in append mode
        existing_processed_ids = set()
        if self.append_mode:
            existing_raw, _ = self.load_existing_processed_data()
            existing_processed_ids = self.get_processed_patient_ids(existing_raw)
        
        try:
            # Read CSV with different encodings
            df = None
            for encoding in ['utf-8', 'latin-1', 'cp1252']:
                try:
                    df = pd.read_csv(self.eipl_csv_file, encoding=encoding)
                    logger.info(f"[LOAD] Loaded {len(df)} records with {encoding} encoding")
                    break
                except UnicodeDecodeError:
                    continue
            
            if df is None:
                logger.error("[ERROR] Could not read CSV file with any encoding")
                return pd.DataFrame()
            
            # Process records
            processed_records = []
            skipped_count = 0
            
            for idx, row in df.iterrows():
                try:
                    # Skip header rows and invalid rows
                    sr_no = row.get('Sr. No.')
                    if pd.isna(sr_no) or str(sr_no).strip() == '' or str(sr_no).strip() == 'Sr. No.':
                        continue
                    
                    # Check if it's a valid numeric Sr. No.
                    try:
                        sr_no_num = int(float(str(sr_no)))
                        if sr_no_num <= 0:
                            continue
                    except (ValueError, TypeError):
                        continue
                    
                    # Extract patient data
                    patient_id = str(row.get('Patient ID', f'EIPL_{idx}')).strip()
                    if patient_id == 'nan' or patient_id == '':
                        patient_id = f'EIPL_{idx}'
                    
                    # Skip if already processed (in append mode)
                    if self.append_mode and patient_id in existing_processed_ids:
                        logger.debug(f"[SKIP] Patient {patient_id} already processed")
                        skipped_count += 1
                        continue
                    
                    # Get hemoglobin reading - try different column names
                    hb_reading = None
                    
                    # Try BIO-CHEQ column (with or without space)
                    if 'BIO-CHEQ' in row and pd.notna(row['BIO-CHEQ']):
                        hb_reading = row['BIO-CHEQ']
                    elif 'BIO-CHEQ ' in row and pd.notna(row['BIO-CHEQ ']):
                        hb_reading = row['BIO-CHEQ ']
                    elif 'Device value (in conc.)' in row and pd.notna(row['Device value (in conc.)']):
                        hb_reading = row['Device value (in conc.)']
                    
                    # Skip if no valid reading
                    if hb_reading is None or pd.isna(hb_reading) or str(hb_reading).strip() == '':
                        continue
                    
                    # Clean up reading
                    hb_reading_str = str(hb_reading).strip()
                    
                    # Extract numeric value
                    numeric_match = re.search(r'(\d+\.?\d*)', hb_reading_str)
                    if not numeric_match:
                        continue
                    
                    hb_value = float(numeric_match.group(1))
                    
                    # Add units if not present
                    if not any(unit in hb_reading_str.lower() for unit in ['g/dl', 'mg/dl']):
                        hb_reading_str = f"{hb_value} g/dL"
                    
                    # Get demographics
                    age = row.get('Age', 35)
                    try:
                        age = int(float(age)) if pd.notna(age) and str(age).upper() != 'NA' else 35
                    except (ValueError, TypeError):
                        age = 35
                    
                    gender = str(row.get('Gender', 'Male')).strip()
                    if gender.lower() in ['unknown', 'nan', '', 'na']:
                        gender = 'Male'  # Default
                    
                    # Get gold standard if available
                    gold_standard = row.get('Gold Sandard', row.get('Gold Standard', ''))
                    if pd.isna(gold_standard):
                        gold_standard = ''
                    
                    # Create record
                    record = {
                        'ID': patient_id,
                        'Test_name': 'Hemoglobin',
                        'reading': hb_reading_str,
                        'numeric_value': hb_value,
                        'Age': age,
                        'Gender': gender,
                        'Gold_Standard': gold_standard,
                        'Source_File': 'EIPL_BIO-CHEQ',
                        'Timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                    }
                    
                    processed_records.append(record)
                    
                except Exception as e:
                    logger.warning(f"[WARNING] Error processing row {idx}: {e}")
                    continue
            
            if not processed_records:
                if self.append_mode and skipped_count > 0:
                    logger.info(f"[APPEND] No new records found - {skipped_count} patients already processed")
                    print(f"[INFO] No new hemoglobin data found - {skipped_count} patients already in Excel file")
                else:
                    logger.warning("[WARNING] No valid hemoglobin records found")
                return pd.DataFrame()
            
            result_df = pd.DataFrame(processed_records)
            if self.append_mode:
                logger.info(f"[APPEND] Found {len(result_df)} NEW hemoglobin records (skipped {skipped_count} existing)")
                print(f"[APPEND] Found {len(result_df)} new patients (skipped {skipped_count} existing)")
            else:
                logger.info(f"[SUCCESS] Processed {len(result_df)} hemoglobin records")
            
            return result_df
            
        except Exception as e:
            logger.error(f"[ERROR] Error processing EIPL CSV: {e}")
            return pd.DataFrame()
    
    def classify_hemoglobin(self, hb_value, gender, age=None):
        """Classify hemoglobin result based on clinical ranges"""
        
        # Normalize gender
        gender_key = 'Male'
        if str(gender).upper().startswith('F'):
            gender_key = 'Female'
        
        ranges = HB_CLINICAL_RANGES[gender_key]
        
        # Simple classification logic:
        # < 10: Severe Anemia (RED)
        # 10 to lower normal limit: Mild Anemia (YELLOW) 
        # Normal range: Normal (GREEN)
        # > upper normal limit: Hemoglobinemia (BLUE)
        
        normal_min, normal_max = ranges['Normal']
        
        if hb_value < 10.0:
            return 'Severe Anemia', ranges['Severe Anemia']
        elif hb_value < normal_min:
            return 'Mild Anemia', ranges['Mild Anemia']
        elif hb_value <= normal_max:
            return 'Normal', ranges['Normal']
        else:
            return 'Hemoglobinemia', ranges['Hemoglobinemia']
    
    def analyze_hemoglobin_data(self, raw_data):
        """Analyze hemoglobin data and create classifications"""
        if raw_data.empty:
            logger.warning("[WARNING] No data to analyze")
            return pd.DataFrame()
        
        logger.info(f"[ANALYZE] Analyzing {len(raw_data)} hemoglobin records")
        
        analyzed_records = []
        
        for _, row in raw_data.iterrows():
            try:
                hb_value = row['numeric_value']
                age = row['Age']
                gender = row['Gender']
                
                # Classify
                classification, classification_range = self.classify_hemoglobin(hb_value, gender, age)
                
                # Get the actual normal range for this patient's gender
                gender_key = 'Male' if not str(gender).upper().startswith('F') else 'Female'
                actual_normal_range = HB_CLINICAL_RANGES[gender_key]['Normal']
                
                # Calculate deviation
                range_mid = (actual_normal_range[0] + actual_normal_range[1]) / 2
                deviation = ((hb_value - range_mid) / range_mid) * 100
                
                # Get risk info
                risk_info = RISK_COLORS[classification]
                
                # Create analyzed record (simplified)
                analyzed_record = {
                    'Patient_ID': row['ID'],
                    'Test_Name': row['Test_name'],
                    'Age': age,
                    'Gender': gender,
                    'Hemoglobin_Value': f"{hb_value:.1f} g/dL",
                    'Classification': classification,
                    'Normal_Range': f"{actual_normal_range[0]:.1f}-{actual_normal_range[1]:.1f} g/dL",
                    'Color_Code': risk_info['color']
                }
                
                # Add gold standard comparison if available (optional)
                if row['Gold_Standard'] and str(row['Gold_Standard']).strip() != '':
                    try:
                        gold_val = float(row['Gold_Standard'])
                        accuracy_error = abs(hb_value - gold_val)
                        analyzed_record['Gold_Standard'] = f"{gold_val:.1f} g/dL"
                        analyzed_record['Accuracy_Error'] = f"{accuracy_error:.2f} g/dL"
                    except (ValueError, TypeError):
                        pass
                
                analyzed_records.append(analyzed_record)
                
            except Exception as e:
                logger.error(f"[ERROR] Error analyzing record {row['ID']}: {e}")
                continue
        
        analyzed_df = pd.DataFrame(analyzed_records)
        logger.info(f"[SUCCESS] Analyzed {len(analyzed_df)} records")
        
        return analyzed_df
    
    def display_results(self, analyzed_df):
        """Display hemoglobin analysis results"""
        if analyzed_df.empty:
            print("[INFO] No results to display")
            return
        
        print("\n" + "="*80)
        print("[REPORT] HEMOGLOBIN ANALYSIS RESULTS")
        print("="*80)
        
        # Sort by classification severity
        classification_order = {'Normal': 1, 'Mild Anemia': 2, 'Severe Anemia': 3, 'Hemoglobinemia': 4}
        analyzed_df['sort_order'] = analyzed_df['Classification'].map(classification_order).fillna(2)
        sorted_results = analyzed_df.sort_values('sort_order', ascending=False)
        
        # Display individual results
        print("\n[INDIVIDUAL RESULTS]")
        print("-" * 50)
        
        for _, row in sorted_results.iterrows():
            print(f"\n{RISK_COLORS[row['Classification']]['symbol']} {row['Color_Code']} - {row['Classification']}")
            print(f"   [ID] Patient: {row['Patient_ID']}")
            print(f"   [TEST] Test: {row['Test_Name']}")
            print(f"   [HB] Hemoglobin: {row['Hemoglobin_Value']}")
            print(f"   [NORMAL] Range: {row['Normal_Range']}")
            print(f"   [DEMO] Age: {row['Age']}, Gender: {row['Gender']}")
            
            # Show gold standard comparison if available
            if 'Gold_Standard' in row and pd.notna(row['Gold_Standard']):
                print(f"   [GOLD] Gold Standard: {row['Gold_Standard']}")
                print(f"   [ACCURACY] Error: {row.get('Accuracy_Error', 'N/A')}")
        
        # Summary statistics
        print("\n" + "="*80)
        print("[SUMMARY] HEMOGLOBIN CLASSIFICATION SUMMARY")
        print("="*80)
        
        total = len(analyzed_df)
        classification_counts = analyzed_df['Classification'].value_counts()
        
        print(f"\nTotal Patients Analyzed: {total}")
        print("\nClassification Distribution:")
        
        for classification in ['Normal', 'Mild Anemia', 'Severe Anemia', 'Hemoglobinemia']:
            count = classification_counts.get(classification, 0)
            percentage = (count / total * 100) if total > 0 else 0
            risk_info = RISK_COLORS[classification]
            print(f"   {risk_info['symbol']} {risk_info['color']} {classification}: {count} patients ({percentage:.1f}%)")
        
        # Gender analysis
        print("\n[GENDER ANALYSIS]")
        gender_groups = analyzed_df.groupby('Gender')
        for gender, group_data in gender_groups:
            abnormal_count = len(group_data[group_data['Classification'] != 'Normal'])
            total_count = len(group_data)
            percentage = (abnormal_count / total_count * 100) if total_count > 0 else 0
            print(f"   {gender}: {abnormal_count}/{total_count} with anemia ({percentage:.1f}%)")
        
        # Age group analysis
        age_groups = {
            'Pediatric (0-12)': analyzed_df[analyzed_df['Age'] <= 12],
            'Adolescent (13-18)': analyzed_df[(analyzed_df['Age'] > 12) & (analyzed_df['Age'] <= 18)],
            'Adult (19-65)': analyzed_df[(analyzed_df['Age'] > 18) & (analyzed_df['Age'] <= 65)],
            'Elderly (65+)': analyzed_df[analyzed_df['Age'] > 65]
        }
        
        print("\n[AGE GROUP ANALYSIS]")
        for group_name, group_data in age_groups.items():
            if not group_data.empty:
                abnormal_count = len(group_data[group_data['Classification'] != 'Normal'])
                total_count = len(group_data)
                percentage = (abnormal_count / total_count * 100) if total_count > 0 else 0
                print(f"   {group_name}: {abnormal_count}/{total_count} with anemia ({percentage:.1f}%)")
        
        # Critical alerts
        severe_cases = analyzed_df[analyzed_df['Classification'] == 'Severe Anemia']
        if not severe_cases.empty:
            print(f"\n🚨 CRITICAL ALERT: {len(severe_cases)} patients with SEVERE ANEMIA require IMMEDIATE attention!")
            for _, case in severe_cases.iterrows():
                print(f"   🔴 Patient {case['Patient_ID']}: {case['Hemoglobin_Value']} (Normal: {case['Normal_Range']})")

        print("\n" + "="*80)
    
    def display_analytical_report(self, analyzed_df, stats):
        """Display comprehensive analytical report"""
        print("\n" + "="*80)
        print("[ANALYTICAL REPORT] STATISTICAL ANALYSIS")
        print("="*80)
        
        # Basic Statistics
        print(f"\n[DESCRIPTIVE STATISTICS]")
        print(f"Total Patients: {stats['total_patients']}")
        print(f"Mean Hemoglobin: {stats['mean_hb']:.2f} g/dL")
        print(f"Median Hemoglobin: {stats['median_hb']:.2f} g/dL")
        print(f"Standard Deviation: {stats['std_hb']:.2f} g/dL")
        print(f"Range: {stats['min_hb']:.1f} - {stats['max_hb']:.1f} g/dL")
        print(f"Coefficient of Variation: {stats['coefficient_of_variation']:.1f}%")
        
        # Percentiles
        print(f"\n[PERCENTILE ANALYSIS]")
        percentiles = stats['percentiles']
        print(f"5th Percentile: {percentiles['5th']:.2f} g/dL")
        print(f"25th Percentile (Q1): {percentiles['25th']:.2f} g/dL")
        print(f"75th Percentile (Q3): {percentiles['75th']:.2f} g/dL")
        print(f"95th Percentile: {percentiles['95th']:.2f} g/dL")
        
        # Classification Distribution
        print(f"\n[CLASSIFICATION DISTRIBUTION]")
        class_pct = stats['classification_percentages']
        print(f"Normal: {class_pct['Normal']:.1f}%")
        print(f"Mild Anemia: {class_pct['Mild_Anemia']:.1f}%")
        print(f"Severe Anemia: {class_pct['Severe_Anemia']:.1f}%")
        print(f"Hemoglobinemia: {class_pct['Hemoglobinemia']:.1f}%")
        
        # Gender Analysis
        if 'male_stats' in stats and 'female_stats' in stats:
            print(f"\n[GENDER COMPARISON]")
            male = stats['male_stats']
            female = stats['female_stats']
            print(f"Males: {male['count']} patients, Mean Hb: {male['mean_hb']:.2f} g/dL, Anemia Rate: {male['anemia_rate']:.1f}%")
            print(f"Females: {female['count']} patients, Mean Hb: {female['mean_hb']:.2f} g/dL, Anemia Rate: {female['anemia_rate']:.1f}%")
            
            # Statistical significance test (simplified)
            hb_diff = abs(male['mean_hb'] - female['mean_hb'])
            print(f"Mean Hemoglobin Difference: {hb_diff:.2f} g/dL")
        
        # Age Group Analysis
        if 'age_group_analysis' in stats:
            print(f"\n[AGE GROUP ANALYSIS]")
            for group, data in stats['age_group_analysis'].items():
                print(f"{group}: {data['count']} patients, Mean Hb: {data['mean_hb']:.2f} g/dL, Anemia Rate: {data['anemia_rate']:.1f}%")
        
        # Outlier Analysis
        print(f"\n[OUTLIER DETECTION]")
        outliers = stats['outliers']
        print(f"Outliers Detected: {outliers['count']} ({outliers['percentage']:.1f}%)")
        if outliers['count'] > 0:
            print(f"Outlier Patient Indices: {outliers['indices']}")
        
        # Correlation Analysis
        if 'age_hb_correlation' in stats:
            print(f"\n[CORRELATION ANALYSIS]")
            corr = stats['age_hb_correlation']
            print(f"Age-Hemoglobin Correlation: {corr['coefficient']:.3f} ({corr['strength']} {corr['direction']})")
        
        # Risk Stratification
        print(f"\n[RISK STRATIFICATION]")
        risk = stats['risk_stratification']
        print(f"🔴 High Risk: {risk['high_risk']['count']} patients ({risk['high_risk']['percentage']:.1f}%)")
        print(f"🟡 Medium Risk: {risk['medium_risk']['count']} patients ({risk['medium_risk']['percentage']:.1f}%)")
        print(f"🟢 Low Risk: {risk['low_risk']['count']} patients ({risk['low_risk']['percentage']:.1f}%)")
        
        print("\n" + "="*80)
    
    def save_excel_with_colors(self, analyzed_data):
        """Save analyzed data to Excel file with colored cells"""
        if not EXCEL_AVAILABLE:
            logger.warning("[WARNING] openpyxl not available - skipping Excel export")
            return False
        
        if analyzed_data.empty:
            logger.warning("[WARNING] No data to save to Excel")
            return False
        
        try:
            excel_file = "hb_analyzed_results.xlsx"
            
            # Handle append mode - load existing Excel if it exists
            if self.append_mode and os.path.exists(excel_file):
                logger.info("[APPEND] Loading existing Excel file for append mode")
                try:
                    # Load existing Excel file
                    from openpyxl import load_workbook
                    wb = load_workbook(excel_file)
                    
                    # Get existing data worksheet
                    if "Hemoglobin Analysis" in wb.sheetnames:
                        ws = wb["Hemoglobin Analysis"]
                        
                        # Read existing data (skip the first 4 rows which are headers/title)
                        existing_data = []
                        for row in ws.iter_rows(min_row=5, values_only=True):
                            if row[0] is not None:  # If first column has data
                                existing_data.append(row)
                        
                        if existing_data:
                            # Convert to DataFrame to combine with new data
                            existing_df = pd.DataFrame(existing_data, columns=[
                                'Patient_ID', 'Test_Name', 'Age', 'Gender', 'Hemoglobin_Value', 
                                'Classification', 'Normal_Range', 'Color_Code'
                            ])
                            
                            # Add gold standard columns if they exist
                            if len(existing_data[0]) > 8:
                                existing_df = pd.DataFrame(existing_data, columns=[
                                    'Patient_ID', 'Test_Name', 'Age', 'Gender', 'Hemoglobin_Value', 
                                    'Classification', 'Normal_Range', 'Color_Code', 'Gold_Standard', 'Accuracy_Error'
                                ])
                            
                            # Combine existing and new data
                            combined_data = pd.concat([existing_df, analyzed_data], ignore_index=True)
                            # Remove duplicates based on Patient_ID (keep the new one)
                            combined_data = combined_data.drop_duplicates(subset=['Patient_ID'], keep='last')
                            analyzed_data = combined_data
                            logger.info(f"[APPEND] Combined {len(existing_df)} existing + {len(analyzed_data) - len(existing_df)} new records")
                    
                    # Remove the old worksheet and create fresh one with combined data
                    if "Hemoglobin Analysis" in wb.sheetnames:
                        wb.remove(wb["Hemoglobin Analysis"])
                    
                except Exception as e:
                    logger.warning(f"[WARNING] Could not load existing Excel for append: {e}")
                    logger.info("[FALLBACK] Creating new Excel file instead")
                    wb = Workbook()
            else:
                # Create new workbook
                wb = Workbook()
            
            # Create/recreate the main data worksheet
            ws = wb.create_sheet("Hemoglobin Analysis")
            if "Sheet" in wb.sheetnames:
                wb.remove(wb["Sheet"])
            wb.active = ws
            
            # Remove the sort_order column if it exists for Excel export
            excel_data = analyzed_data.copy()
            if 'sort_order' in excel_data.columns:
                excel_data = excel_data.drop('sort_order', axis=1)
            
            # Write data to worksheet
            for r in dataframe_to_rows(excel_data, index=False, header=True):
                ws.append(r)
            
            # Style the header row
            header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            header_font = Font(color='FFFFFF', bold=True)
            header_alignment = Alignment(horizontal='center', vertical='center')
            
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment
            
            # Apply colors based on classification
            classification_col = None
            color_code_col = None
            
            # Find the column indices
            for idx, cell in enumerate(ws[1], 1):
                if cell.value == 'Classification':
                    classification_col = idx
                elif cell.value == 'Color_Code':
                    color_code_col = idx
            
            # Apply colors to data rows
            for row in range(2, ws.max_row + 1):
                if classification_col:
                    classification = ws.cell(row=row, column=classification_col).value
                    if classification in EXCEL_COLORS:
                        color_info = EXCEL_COLORS[classification]
                        
                        # Color the entire row
                        for col in range(1, ws.max_column + 1):
                            cell = ws.cell(row=row, column=col)
                            cell.fill = PatternFill(start_color=color_info['fill'], 
                                                  end_color=color_info['fill'], 
                                                  fill_type='solid')
                            cell.font = Font(color=color_info['font'], bold=False)
                            cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Add borders
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            for row in ws.iter_rows():
                for cell in row:
                    cell.border = thin_border
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                
                adjusted_width = min(max_length + 2, 20)  # Cap at 20 characters
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Add a title and summary at the top
            ws.insert_rows(1, 3)
            
            # Title
            ws['A1'] = 'HEMOGLOBIN ANALYSIS REPORT'
            ws['A1'].font = Font(size=16, bold=True, color='2F4F4F')
            ws['A1'].alignment = Alignment(horizontal='center')
            ws.merge_cells('A1:H1')
            
            # Timestamp
            ws['A2'] = f'Generated: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}'
            ws['A2'].font = Font(size=11, italic=True, color='696969')
            ws['A2'].alignment = Alignment(horizontal='center')
            ws.merge_cells('A2:H2')
            
            # Summary
            total_patients = len(analyzed_data)
            normal_count = len(analyzed_data[analyzed_data['Classification'] == 'Normal'])
            anemia_count = len(analyzed_data[analyzed_data['Classification'].isin(['Mild Anemia', 'Severe Anemia'])])
            
            ws['A3'] = f'Total Patients: {total_patients} | Normal: {normal_count} | Anemia Cases: {anemia_count}'
            ws['A3'].font = Font(size=11, bold=True, color='2F4F4F')
            ws['A3'].alignment = Alignment(horizontal='center')
            ws.merge_cells('A3:H3')
            
            # Protect the worksheet with password while allowing sort and filter
            ws.protection.password = "eipl"
            ws.protection.sheet = True
            ws.protection.sort = False  # Allow sorting
            ws.protection.autoFilter = False  # Allow filtering
            ws.protection.selectLockedCells = True  # Allow selecting cells
            ws.protection.selectUnlockedCells = True  # Allow selecting unlocked cells
            
            # Enable AutoFilter for the data range (starting from row 4 which has headers after our title rows)
            if ws.max_row > 4:  # Ensure we have data rows
                ws.auto_filter.ref = f"A4:{chr(64 + ws.max_column)}{ws.max_row}"
            
            # Create/Update Summary Report worksheet
            if "Summary Report" in wb.sheetnames:
                wb.remove(wb["Summary Report"])
            summary_ws = wb.create_sheet("Summary Report")
            self._create_summary_worksheet(summary_ws, analyzed_data)
            
            # Save the workbook
            wb.save(excel_file)
            
            if self.append_mode:
                logger.info(f"[SAVE] Excel file updated with new data (Protected with password)")
                print("   [APPEND] ✨ New data added to existing Excel file!")
            else:
                logger.info(f"[SAVE] Excel file with colors and summary report saved to {excel_file} (Protected with password)")
                print("   [SUCCESS] ✨ Excel file with colored data and summary report created!")
                
            return True
            
        except Exception as e:
            logger.error(f"[ERROR] Error creating Excel file: {e}")
            return False
    
    def _create_summary_worksheet(self, ws, analyzed_data):
        """Create a comprehensive summary report worksheet"""
        try:
            # Title
            ws['A1'] = 'HEMOGLOBIN ANALYSIS SUMMARY REPORT'
            ws['A1'].font = Font(size=18, bold=True, color='2F4F4F')
            ws['A1'].alignment = Alignment(horizontal='center')
            ws.merge_cells('A1:D1')
            
            # Timestamp and mode
            ws['A3'] = f'Generated: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}'
            ws['A3'].font = Font(size=11, italic=True, color='696969')
            
            mode_text = "Mode: APPEND - Added new records" if self.append_mode else "Mode: COMPLETE ANALYSIS"
            ws['A4'] = mode_text
            ws['A4'].font = Font(size=11, color='696969')
            
            # Basic statistics
            total = len(analyzed_data)
            ws['A6'] = 'BASIC STATISTICS'
            ws['A6'].font = Font(size=14, bold=True, color='2F4F4F')
            
            ws['A7'] = f'Total Patients: {total}'
            ws['A7'].font = Font(size=11, bold=True)
            
            # Classification distribution
            ws['A9'] = 'CLASSIFICATION DISTRIBUTION'
            ws['A9'].font = Font(size=14, bold=True, color='2F4F4F')
            
            classification_counts = analyzed_data['Classification'].value_counts()
            row = 10
            for classification in ['Normal', 'Mild Anemia', 'Severe Anemia', 'Hemoglobinemia']:
                count = classification_counts.get(classification, 0)
                percentage = (count / total * 100) if total > 0 else 0
                risk_info = RISK_COLORS[classification]
                
                # Classification name
                ws[f'A{row}'] = f'{risk_info["symbol"]} {classification}:'
                ws[f'A{row}'].font = Font(size=11, bold=True)
                
                # Count and percentage
                ws[f'B{row}'] = f'{count} patients ({percentage:.1f}%)'
                ws[f'B{row}'].font = Font(size=11)
                
                # Color coding
                if classification in EXCEL_COLORS:
                    color_info = EXCEL_COLORS[classification]
                    ws[f'A{row}'].fill = PatternFill(start_color=color_info['fill'], 
                                                   end_color=color_info['fill'], 
                                                   fill_type='solid')
                    ws[f'B{row}'].fill = PatternFill(start_color=color_info['fill'], 
                                                   end_color=color_info['fill'], 
                                                   fill_type='solid')
                
                row += 1
            
            # Gender analysis
            row += 1
            ws[f'A{row}'] = 'GENDER ANALYSIS'
            ws[f'A{row}'].font = Font(size=14, bold=True, color='2F4F4F')
            row += 1
            
            gender_groups = analyzed_data.groupby('Gender')
            for gender, group_data in gender_groups:
                abnormal_count = len(group_data[group_data['Classification'] != 'Normal'])
                total_count = len(group_data)
                percentage = (abnormal_count / total_count * 100) if total_count > 0 else 0
                
                ws[f'A{row}'] = f'{gender}:'
                ws[f'A{row}'].font = Font(size=11, bold=True)
                ws[f'B{row}'] = f'{abnormal_count}/{total_count} with anemia ({percentage:.1f}%)'
                ws[f'B{row}'].font = Font(size=11)
                row += 1
            
            # Age group analysis
            row += 1
            ws[f'A{row}'] = 'AGE GROUP ANALYSIS'
            ws[f'A{row}'].font = Font(size=14, bold=True, color='2F4F4F')
            row += 1
            
            age_groups = {
                'Pediatric (0-12)': analyzed_data[analyzed_data['Age'] <= 12],
                'Adolescent (13-18)': analyzed_data[(analyzed_data['Age'] > 12) & (analyzed_data['Age'] <= 18)],
                'Adult (19-65)': analyzed_data[(analyzed_data['Age'] > 18) & (analyzed_data['Age'] <= 65)],
                'Elderly (65+)': analyzed_data[analyzed_data['Age'] > 65]
            }
            
            for group_name, group_data in age_groups.items():
                if not group_data.empty:
                    abnormal_count = len(group_data[group_data['Classification'] != 'Normal'])
                    total_count = len(group_data)
                    percentage = (abnormal_count / total_count * 100) if total_count > 0 else 0
                    
                    ws[f'A{row}'] = f'{group_name}:'
                    ws[f'A{row}'].font = Font(size=11, bold=True)
                    ws[f'B{row}'] = f'{abnormal_count}/{total_count} with anemia ({percentage:.1f}%)'
                    ws[f'B{row}'].font = Font(size=11)
                    row += 1
            
            # Critical alerts
            severe_cases = analyzed_data[analyzed_data['Classification'] == 'Severe Anemia']
            if not severe_cases.empty:
                row += 1
                ws[f'A{row}'] = 'CRITICAL ALERTS'
                ws[f'A{row}'].font = Font(size=14, bold=True, color='8B0000')  # Dark red
                row += 1
                
                ws[f'A{row}'] = f'🚨 URGENT: {len(severe_cases)} patients with SEVERE ANEMIA!'
                ws[f'A{row}'].font = Font(size=12, bold=True, color='8B0000')
                ws[f'A{row}'].fill = PatternFill(start_color='FFB6C1', end_color='FFB6C1', fill_type='solid')
                row += 1
                
                for _, case in severe_cases.iterrows():
                    ws[f'A{row}'] = f'🔴 Patient {case["Patient_ID"]}:'
                    ws[f'A{row}'].font = Font(size=10, bold=True)
                    ws[f'B{row}'] = f'{case["Hemoglobin_Value"]} (Normal: {case["Normal_Range"]})'
                    ws[f'B{row}'].font = Font(size=10)
                    ws[f'A{row}'].fill = PatternFill(start_color='FFB6C1', end_color='FFB6C1', fill_type='solid')
                    ws[f'B{row}'].fill = PatternFill(start_color='FFB6C1', end_color='FFB6C1', fill_type='solid')
                    row += 1
            
            # Auto-adjust column widths
            ws.column_dimensions['A'].width = 35
            ws.column_dimensions['B'].width = 30
            ws.column_dimensions['C'].width = 20
            ws.column_dimensions['D'].width = 20
            
            # Apply protection to summary worksheet
            ws.protection.password = "eipl"
            ws.protection.sheet = True
            ws.protection.selectLockedCells = True
            ws.protection.selectUnlockedCells = True
            
            logger.info("[SUMMARY] Summary worksheet created successfully")
            
        except Exception as e:
            logger.error(f"[ERROR] Error creating summary worksheet: {e}")
    
    def save_results(self, raw_data, analyzed_data):
        """Save results to Excel with colors and integrated summary (no separate files)"""
        try:
            if not analyzed_data.empty:
                # Only create Excel with colors and integrated summary
                excel_success = self.save_excel_with_colors(analyzed_data)
                if excel_success:
                    if not self.append_mode:
                        print("   [SUCCESS] ✨ Excel file with colored data and summary report created!")
                    print("   [INFO] 📊 Two worksheets: 'Hemoglobin Analysis' (data) + 'Summary Report' (analysis)")
                    logger.info(f"[SAVE] Complete Excel analysis saved with {len(analyzed_data)} records and summary")
            
        except Exception as e:
            logger.error(f"[ERROR] Error saving results: {e}")
    
    def run_complete_analysis(self):
        """Run complete hemoglobin analysis workflow"""
        print("\n" + "="*80)
        print("[SYSTEM] HEMOGLOBIN MEDICAL DATA PROCESSOR")
        print("="*80)
        print(f"[START] Starting analysis at {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        
        try:
            # Step 1: Process EIPL CSV
            print("\n[STEP 1] Processing EIPL CSV file...")
            raw_data = self.process_eipl_csv()
            
            if raw_data.empty:
                print("[ERROR] No data found to process!")
                return False
            
            print(f"[SUCCESS] Processed {len(raw_data)} hemoglobin records")
            
            # Step 2: Analyze data
            print("\n[STEP 2] Analyzing hemoglobin data...")
            analyzed_data = self.analyze_hemoglobin_data(raw_data)
            
            if analyzed_data.empty:
                print("[ERROR] Analysis failed!")
                return False
            
            print(f"[SUCCESS] Analyzed {len(analyzed_data)} records")
            
            # Step 3: Display results
            print("\n[STEP 3] Displaying results...")
            self.display_results(analyzed_data)
            
            # Step 4: Save results
            print("\n[STEP 4] Saving results...")
            self.save_results(raw_data, analyzed_data)
            
            print("\n" + "="*80)
            print("[COMPLETE] Hemoglobin analysis completed successfully!")
            print("="*80)
            
            return True
            
        except Exception as e:
            logger.error(f"[ERROR] Analysis failed: {e}")
            print(f"\n[ERROR] Analysis failed: {e}")
            return False


def main():
    """Main function"""
    try:
        # Check if user wants append mode
        print("\n" + "="*80)
        print("[HEMOGLOBIN PROCESSOR] Smart Excel Analysis")
        print("="*80)
        print("[INFO] This processor generates ONE Excel file with:")
        print("[INFO] • 'Hemoglobin Analysis' sheet - Color-coded patient data")
        print("[INFO] • 'Summary Report' sheet - Complete analysis summary")
        print("[INFO] • Password protection with sorting/filtering enabled")
        print("="*80)
        print("[APPEND FEATURE] You can add new patients to your EIPL CSV")
        print("[APPEND FEATURE] and choose append mode to add them to existing Excel!")
        print("="*80)
        
        choice = input("\nChoose mode:\n[1] Complete analysis (overwrites existing Excel)\n[2] Append new data (adds to existing Excel)\nPress Enter for complete analysis: ").strip()
        
        append_mode = False
        if choice == "2":
            append_mode = True
            print("\n[MODE] APPEND MODE - Adding new patients to existing Excel file")
        else:
            append_mode = False
            print("\n[MODE] COMPLETE MODE - Generating new Excel file")
        
        processor = HemoglobinProcessor(append_mode=append_mode)
        success = processor.run_complete_analysis()
        
        if success:
            print("\n[FILES] Generated file:")
            excel_file = "hb_analyzed_results.xlsx"
            if os.path.exists(excel_file):
                file_size = os.path.getsize(excel_file)
                print(f"   [OK] {excel_file} ({file_size} bytes)")
                print("        ├── 'Hemoglobin Analysis' sheet: Color-coded patient data with filtering")
                print("        └── 'Summary Report' sheet: Complete statistical analysis and recommendations")
        
    except Exception as e:
        print(f"[FATAL] Fatal error: {e}")
    
    print("\nPress Enter to exit...")
    input()


if __name__ == "__main__":
    main()
